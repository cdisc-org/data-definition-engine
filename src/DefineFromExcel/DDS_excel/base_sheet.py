import os
import datetime
import pandas as pd
from openpyxl import load_workbook
from DDS_excel.globals import Globals


class BaseSheet:
    class StateError(Exception):
        pass

    class FormatError(Exception):
        pass

    def __init__(
        self,
        file_path: str,
        globals: Globals,
        sheet_name: str,
        header: int = 0,
        optional: bool = False,
        converters: dict = {},
        require: dict = {},
    ):
        self.file_path = file_path
        self.globals = globals
        self.dir_path, self.filename = os.path.split(file_path)
        self.sheet_name = sheet_name
        self.sheet = pd.DataFrame()
        self.success = False
        self._sheet_names = None
        if optional and not self._sheet_present(file_path, sheet_name):
            self._general_info(f"'{sheet_name}' not found but optional")
        else:
            if require and not self._check_cell_value(
                file_path,
                sheet_name,
                require["row"],
                require["column"],
                require["value"],
            ):
                self._general_info(
                    f"Required value {require['value']} at [{require['row']}, {require['column']}] mismatch in {sheet_name}"
                )
                pass
            else:
                self.sheet = pd.read_excel(
                    open(file_path, "rb"),
                    sheet_name=sheet_name,
                    header=header,
                    converters=converters,
                )
                self.success = True
                self._general_info("Processed sheet %s" % (sheet_name))


    def _general_info(self, message):
        self.globals.errors_and_logging.info(message,self.sheet_name)

    def _get_sheet_names(self, file_path):
        if not self._sheet_names:
            wb = load_workbook(file_path, read_only=True, keep_links=False)
            self._sheet_names = wb.sheetnames
        return self._sheet_names

    def _sheet_exception(self, e):
        self.globals.errors_and_logging.exception(
            f"Error [{e}] while reading sheet '{self.sheet_name}'", e, self.sheet_name
        )


    def _sheet_present(self, file_path, sheet_name):
        sheet_names = self._get_sheet_names(file_path)
        return sheet_name in sheet_names

    def _check_cell_value(self, file_path, sheet_name, row, column, value):
        wb = load_workbook(file_path, read_only=True, keep_links=False)
        ws = wb[sheet_name]
        # print(f"CELL={ws.cell(row, column).value}")
        return str(ws.cell(row, column).value).upper() == value

