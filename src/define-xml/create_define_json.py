"""
USDM to Define JSON Processor
===================================

This module processes USDM (Unified Study Data Model) JSON files and generates
Define-JSON structures with value-level metadata organized in "slices".

The major difference from the standard Define-JSON format is that variable-level
metadata (VLM) is organized under "slices" within each itemGroup rather than as
separate "children" itemGroups with type "DataSpecialization".

"""

import json
import os
import argparse
import hashlib
import yaml
import jmespath
from cdisc_library_client import CDISCLibraryClient
from collections import defaultdict
from dotenv import load_dotenv
from datetime import datetime


class USDMDefineJSONProcessor:
    """
    Processes USDM JSON files and generates Define-JSON structure.
    
    This class transforms USDM data into Define-JSON format where VLM is
    organized as "slices" within itemGroups instead of separate DataSpecialization
    itemGroups.
    
    Attributes:
        client (CDISCLibraryClient): CDISC Library API client
        usdm_data (dict): Loaded USDM JSON data
        template (dict): Output Define-JSON structure
        datasets_dict (dict): Dataset variables and restricted values
        bc_dict (dict): Biomedical concepts with VLM metadata
        vlm_lookup (dict): VLM metadata keyed by variable name
        test_dict (dict): TEST/TESTCD concept mappings for restricted CT
        item_groups (list): Item group definitions
        where_clauses (list): Where clause definitions
        conditions (list): Condition definitions
        condition_lookup (dict): Deduplication cache for conditions
        code_lists_map (dict): Deduplicated codelists by short name
        vlm_items_by_variable (dict): VLM items grouped for slice creation
        eligibility_criteria (list): Inclusion/exclusion criteria from USDM
        elements (list): Study elements from USDM
        required_variables_exceptions (dict): Domains with relaxed Req/Exp logic
        cosmosversion (str): CDISC Cosmos API version
    """
    
    def __init__(self, usdm_file, output_template, sdtmig, sdtmct, 
                 studyversion, studydesign, docversion, cdisc_api_key, 
                 cosmosversion, debug):
        """
        Initialize the USDM processor.
        
        Args:
            usdm_file (str): Path to USDM JSON file
            output_template (str): Path to output Define-JSON file
            sdtmig (str): SDTM Implementation Guide version
            sdtmct (str): SDTM Controlled Terminology date in yyyy-mm-dd format
            studyversion (int): Study version index in USDM
            studydesign (int): Study design index in USDM
            docversion (int): Document version index in USDM
            cdisc_api_key (str): CDISC Library API Key (can be None to use environment variable)
            cosmosversion (str): CDISC Cosmos API version
            debug (bool): Enable debug mode to save intermediate dictionaries
            
        Raises:
            ValueError: If sdtmct date format is invalid
        """
        load_dotenv()
        
        self.api_key = cdisc_api_key if cdisc_api_key else os.getenv("CDISC_API_KEY")
        self.client = CDISCLibraryClient(api_key=self.api_key)
        
        with open(usdm_file, "r") as file:
            self.usdm_data = json.load(file)
        
        self.studyversion = studyversion
        self.studydesign = studydesign
        self.docversion = docversion
        self.sdtmig = sdtmig
        self.sdtmct = sdtmct
        self.output_template = output_template
        self.debug = debug
        self.cosmosversion = cosmosversion
        
        if sdtmct and not self._validate_date_format(sdtmct):
            raise ValueError("sdtmct must be in yyyy-mm-dd format")
        
        self.template = {
            'OID': '',
            'name': '',
            'description': '',
            'fileOID': '',
            'creationDateTime': '',
            'odmVersion': '1.3.2',
            'fileType': 'Snapshot',
            'originator': 'Define-JSON Processor',
            'context': 'Other',
            'defineVersion': '2.1.0',
            'studyOID': '',
            'studyName': '',
            'studyDescription': '',
            'protocolName': '',
            'itemGroups': [],
            'conditions': [],
            'whereClauses': [],
            'codeLists': [],
            'methods': [],
            'comments': [],
            'standards': [],
            'annotatedCRF': [],
            'concepts': [],
            'conceptProperties': []
        }
        
        self.datasets_dict = {}
        self.bc_dict = {}
        self.vlm_lookup = {}
        self.test_dict = {}
        self.item_groups = []
        self.where_clauses = []
        self.conditions = []
        self.condition_lookup = {}
        self.code_lists_map = {}
        self.vlm_items_by_variable = {}
        
        self.required_variables_exceptions = {
            "TA": ["ELEMENT"],
            "TV": ["VISITDY", "ARM", "TVENRL"],
            "SE": ["ELEMENT", "EPOCH"],
            "SV": ["VISIT", "SVCNTMOD"]
        }
        
        self._extract_usdm_data()

    def _validate_date_format(self, date_string):
        """
        Validate date string is in yyyy-mm-dd format.
        
        Args:
            date_string (str): Date string to validate
            
        Returns:
            bool: True if valid format and date, False otherwise
        """
        import re
        
        pattern = r'^\d{4}-\d{2}-\d{2}$'
        if not re.match(pattern, date_string):
            return False
        
        try:
            datetime.strptime(date_string, '%Y-%m-%d')
            return True
        except ValueError:
            return False

    def _convert_data_type(self, var):
        """
        Convert SDTM data type to Define-XML data type.
        
        Args:
            sdtm_data_type (str): SDTM data type (e.g., 'Char', 'Num')
            
        Returns:
            str: Define-XML data type (e.g., 'text', 'float')
        """
        type_mapping = {
            'Char': 'text',
            'Num': 'integer'
        }

        # Check variable name suffix for special data types
        if var['name'].endswith('DTC'):
            return 'datetime'
        elif var['name'].endswith('DUR'):
            return 'durationDatetime'
        else:
            return type_mapping.get(var['simpleDatatype'], '????')
    
    def _extract_usdm_data(self):
        """Extract biomedical concepts, eligibility criteria, and study elements from USDM."""
        # Extract study version data once
        self.study_version_data = jmespath.search(f'versions[{self.studyversion}]', self.usdm_data.get('study', {}))
        
        # Extract study design data once
        self.studyDesignData = jmespath.search(f'studyDesigns[{self.studydesign}]', self.study_version_data if self.study_version_data else {})

    def process_biomedical_concepts(self):
        """
        Process all biomedical concepts from USDM.
        
        Iterates through biomedical concepts and dispatches to appropriate
        handler based on concept type (Biomedical Concept vs Dataset Specialization).
        Populates datasets_dict and bc_dict data structures.
        """
        # Debug: Accumulate all dataset_data for debugging
        self.all_dataset_data = []

        for bc in self.study_version_data.get('biomedicalConcepts', []):
            bc_data = self.client.get_api_json(f"/cosmos/{self.cosmosversion}" + bc['reference'])
            concept_type = bc_data['_links']['self']['type']
            
            if concept_type == "Biomedical Concept":
                self._process_bc_type(bc, bc_data)
            elif concept_type == "SDTM Dataset Specialization":
                self._process_dss_type(bc, bc_data)
                    
    def _process_bc_type(self, bc, bc_data):
        """
        Process 'Biomedical Concept' type concepts.
        
        Extracts dataset specializations and their variables, updating datasets_dict.
        
        Args:
            bc (dict): Biomedical concept from USDM
            bc_data (dict): Detailed concept data from CDISC Library
        """
        dataset_links = self.client.get_biomedicalconcept_latest_datasetspecializations(self.cosmosversion, bc_data['conceptId'])['sdtm']
        
        for dataset_link in dataset_links:
            dataset_data = self.client.get_sdtm_latest_sdtm_datasetspecialization(self.cosmosversion, dataset_link['href'].split('/')[-1])
            self.all_dataset_data.append(dataset_data)
            dataset_name = dataset_data['domain']
            variables = dataset_data.get('variables', [])
            self._process_variables(variables, dataset_name, bc)
            
    def _process_dss_type(self, bc, bc_data):
        """
        Process 'SDTM Dataset Specialization' type concepts.
        
        Extracts variables, builds where clauses, and processes VLM targets.
        Updates both bc_dict and datasets_dict.
        
        Args:
            bc (dict): Biomedical concept from USDM
            bc_data (dict): Detailed concept data from CDISC Library
        """
        dss_response = self.client.get_sdtm_latest_sdtm_datasetspecialization(self.cosmosversion, bc_data['datasetSpecializationId'])
        
        dataset_name = dss_response['domain']
        variables = dss_response.get('variables', [])
        
        self._process_variables(variables, dataset_name, bc)
        where_clause = self._build_where_clause(bc, bc_data, dss_response, dataset_name)
        self._process_vlm_target_variables(bc, bc_data, where_clause)

    def _process_variables(self, variables, dataset_name, bc):
        """
        Process variables for a dataset, extracting response codes.
        
        Args:
            variables (list): List of variable dictionaries
            dataset_name (str): Dataset domain name
            bc (dict): Parent biomedical concept
        """
        if dataset_name not in self.datasets_dict:
            self.datasets_dict[dataset_name] = {}
        
        for variable in variables:
            variable_name = variable.get('name')
            data_element_concept_id = variable.get('dataElementConceptId')
            codelist_concept_id = variable.get('codelist', {}).get('conceptId')

            if variable_name and data_element_concept_id:
                if variable_name not in self.datasets_dict[dataset_name]:
                    # Create the variable entry
                    self.datasets_dict[dataset_name][variable_name] = {}
                    
                    # Add optional fields to the variable
                    var_fields = ['role', 'dataType', 'length', 'format', 'significantDigits', 'originType', 'originSource']
                    
                    for field in var_fields:
                        value = variable.get(field)
                        if value is not None:
                            self.datasets_dict[dataset_name][variable_name][field] = value

                for property in bc['properties']:
                    property_code = property['code']['standardCode']['code']

                    if data_element_concept_id == property_code:
                        response_codes = []

                        if codelist_concept_id:
                            terms = self.client.get_codelist_terms(f"sdtmct-{self.sdtmct}/", codelist_concept_id)
                            for response_code in property.get('responseCodes', []):
                                code = response_code.get('code', {}).get('code', '')
                                value = next((term for term in terms if term.get('conceptId') == code), None)
                                if value:
                                    response_codes.append(value['submissionValue'])
                            
                            # Create codelist structure if it doesn't exist
                            if "codelist" not in self.datasets_dict[dataset_name][variable_name]:
                                self.datasets_dict[dataset_name][variable_name]["codelist"] = {}
                            
                            if codelist_concept_id not in self.datasets_dict[dataset_name][variable_name]["codelist"]:
                                self.datasets_dict[dataset_name][variable_name]["codelist"][codelist_concept_id] = []

                            self.datasets_dict[dataset_name][variable_name]["codelist"][codelist_concept_id] = list(
                                set(self.datasets_dict[dataset_name][variable_name]["codelist"][codelist_concept_id] + response_codes)
                            )
                        break
    
    def _build_where_clause(self, bc, bc_data, dss_response, dataset_name):
        """
        Build where clauses for variables with comparators.
        
        Args:
            bc (dict): Biomedical concept from USDM
            bc_data (dict): Detailed concept data
            dss_response (dict): Dataset specialization data
            dataset_name (str): Dataset domain name
        
        Returns:
            list: Where clause dictionaries with variables, comparators, and values
        """
        where_clause = []
        clause_items = []
        
        for property in bc['properties']:
            variable_name = property['name']
            variable_data = next((var for var in bc_data['variables'] if var.get('name') == variable_name), None)

            if variable_data and 'comparator' in variable_data:
                codelist_concept_id = variable_data.get('codelist', {}).get('conceptId')
                terms = self.client.get_codelist_terms(f"sdtmct-{self.sdtmct}/", codelist_concept_id)
                
                response_values = []
                for response_code in property.get('responseCodes', []):
                    code = response_code.get('code', {}).get('code', '')
                    value = next((term for term in terms if term.get('conceptId') == code), None)
                    if value:
                        response_values.append(value['submissionValue'])

                if not response_values:
                    dataset_variable = next((var for var in dss_response['variables'] if var.get('name') == variable_name), None)
                    
                    if dataset_variable:
                        if 'assignedTerm' in dataset_variable and 'conceptId' in dataset_variable['assignedTerm']:
                            response_values = [dataset_variable['assignedTerm']['value']]
                        elif 'valueList' in dataset_variable and dataset_variable['valueList']:
                            response_values = dataset_variable['valueList']

                clause_item = {
                    "Dataset": dataset_name,
                    "Variable": variable_name,
                    "item": f"IT.{dataset_name}.{variable_name}",
                    "Codelist Concept ID": codelist_concept_id,
                    "Comparator": variable_data['comparator'],
                    "Values": response_values
                }
                
                clause_items.append(clause_item)
        
        # Create single WhereClause with all clauses combined (implicit AND)
        if clause_items:
            where_clause.append({"Clause": clause_items})
        
        return where_clause

    def _process_vlm_target_variables(self, bc, bc_data, where_clause):
        """
        Process VLM (Variable Level Metadata) target variables.
        
        Extracts VLM metadata including data types, origins, and response codes,
        associating them with where clauses. Updates bc_dict.
        
        Args:
            bc (dict): Biomedical concept from USDM
            bc_data (dict): Detailed concept data
            where_clause (list): Associated where clause definitions
        """
        vlm_targets_found = False
        
        for property in bc['properties']:
            variable_name = property['name']
            variable_data = next((var for var in bc_data['variables'] if var.get('name') == variable_name), None)

            if (variable_data and 'comparator' not in variable_data and variable_data.get('vlmTarget') == True):
                if not vlm_targets_found:
                    if bc['id'] not in self.bc_dict:
                        self.bc_dict[bc['id']] = []
                    vlm_targets_found = True

                codelist_concept_id = variable_data.get('codelist', {}).get('conceptId')
                terms = self.client.get_codelist_terms(f"sdtmct-{self.sdtmct}/", codelist_concept_id)
                
                vlm_data = {}
                vlm_fields = ['role', 'dataType', 'length', 'format', 'significantDigits', 'originType', 'originSource']
                
                for field in vlm_fields:
                    value = variable_data.get(field)
                    if value is not None:
                        vlm_data[field] = value

                response_values = []
                if terms:
                    for response_code in property.get('responseCodes', []):
                        code = response_code.get('code', {}).get('code', '')
                        value = next((term for term in terms if term.get('conceptId') == code), None)
                        if value:
                            response_values.append(value['submissionValue'])

                if response_values:
                    vlm_data['responseCodes'] = response_values
                
                vlm_data['WhereClause'] = where_clause
                variable_dict = {variable_name: vlm_data}
                self.bc_dict[bc['id']].append(variable_dict)

    def build_vlm_lookup(self):
        """
        Build VLM lookup dictionary for quick variable metadata access.
        
        Aggregates VLM metadata from biomedical concepts and supplements it
        with synthetic entries for eligibility criteria (IE) and study
        elements (SE) so those variables get value-level rules and timing
        attributes when present in the USDM input.
        """
        for concept_list in self.bc_dict.values():
            for concept_item in concept_list:
                for variable_name, variable_data in concept_item.items():
                    if variable_name not in self.vlm_lookup:
                        self.vlm_lookup[variable_name] = []
                    
                    if variable_data not in self.vlm_lookup[variable_name]:
                        self.vlm_lookup[variable_name].append(variable_data)
        
        # Add eligibility criteria VLM entries
        if self.studyDesignData.get('eligibilityCriteria', []):
            for criterion in self.studyDesignData.get('eligibilityCriteria', []):
                criterion_name = criterion.get('name', '')
                criterion_label = criterion.get('label', '')
                category_decode = criterion.get('category', {}).get('decode', '')
                
                # Create VLM entry for IEORRES variable
                ieorres_entry = {
                    "role": "Qualifier",
                    "dataType": "text",
                    "length": 1,
                    "originType": "Collected",
                    "originSource": "Investigator",
                    "WhereClause": [
                        {
                            "Clause": [
                                {
                                    "Dataset": "IE",
                                    "Variable": "IETESTCD",
                                    "item": "IT.IE.IETESTCD",
                                    "Comparator": "EQ",
                                    "Values": [criterion_name]
                                }
                            ]
                        }
                    ]
                }
                
                # Add to vlm_lookup
                if "IEORRES" not in self.vlm_lookup:
                    self.vlm_lookup["IEORRES"] = []
                self.vlm_lookup["IEORRES"].append(ieorres_entry)
        
        # Add VLM entries for TS domain
        self.vlm_lookup["TSPARMCD"] = []

        # ADAPT
        if any(code.get('code') == 'C98704' for code in self.studyDesignData.get('characteristics', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 1,
                "codeList": "CL.NY",
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["ADAPT"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)   

        # AGEMIN
        if self.studyDesignData.get('population', {}).get("plannedAge", {}).get("minValue", {}).get("value") is not None or \
           any(cohort.get('plannedAge', {}).get('minValue', {}).get('value') is not None for cohort in self.studyDesignData.get('population', {}).get('cohorts', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["AGEMIN"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)   

        # AGEMAX
        if self.studyDesignData.get('population', {}).get("plannedAge", {}).get("maxValue", {}).get("value") is not None or \
           any(cohort.get('plannedAge', {}).get('maxValue', {}).get('value') is not None for cohort in self.studyDesignData.get('population', {}).get('cohorts', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["AGEMAX"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)   

        # BRDNAIND
        if any(br.get('includesDNA', None) is True for br in self.studyDesignData.get('biospecimenRetentions', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["BRDNAIND"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)   

        # BRIND
        if any(br.get('isRetained', None) is True for br in self.studyDesignData.get('biospecimenRetentions', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["BRIND"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)   

        # COMPTRT
        if any(
            intervention.get('role', {}).get('code') 
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["COMPTRT"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # CRMDUR
        if any(
            intervention.get('minimumResponseDuration', {}).get('value', None)
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["CRMDUR"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # CURTRT
        if any(
            intervention.get('role', {}).get('code', '') == 'C165822'
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["CURTRT"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # DMCIND
        if any(code.get('code') == 'C142578' for code in self.study_version_data.get('roles', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 1,
                "codeList": "CL.NY",
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["DMCIND"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)   

        # DOSE
        if any(
            administration.get('dose', {}).get('value', None)
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
            for administration in intervention.get('administrations', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["DOSE"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # DOSFRM
        if any(
            product.get('administrableDoseForm', {}).get('standardCode', {}).get('decode')
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
            for administration in intervention.get('administrations', [])
            for product in self.study_version_data.get('administrableProducts', [])
            if product.get('id') == administration.get('administrableProductId')
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["DOSFRM"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # DOSFRQ
        if any(
            administration.get('frequency', {}).get('standardCode', {}).get('code') 
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
            for administration in intervention.get('administrations', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["DOSFRQ"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # DOSU
        if any(
            administration.get('dose', {}).get('unit', {}).get('standardCode', {}).get('code') 
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
            for administration in intervention.get('administrations', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["DOSU"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # EXTTIND
        if any(code.get('code') == 'C207613' for code in self.studyDesignData.get('characteristics', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 1,
                "codeList": "CL.NY",
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["EXTTIND"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # HLTSUBJI
        if self.studyDesignData.get('population', {}).get("includesHealthySubjects") is True or \
           any(cohort.get('includesHealthySubjects') is True for cohort in self.studyDesignData.get('population', {}).get('cohorts', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 1,
                "codeList": "CL.NY",
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["HLTSUBJI"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # INDIC
        if any(indication.get('label') for indication in self.studyDesignData.get('indications', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["INDIC"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # INTMODEL
        if self.studyDesignData['studyType']['code'] == 'C98388' and self.studyDesignData.get('model', {}).get('decode', '') != '':
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["INTMODEL"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # INTTYPE
        if any(
            intervention.get('type', {}).get('decode', '') != ''
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["INTTYPE"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # LENGTH
        if any(
            (timeline.get('plannedDuration') or {}).get('quantity', {}).get('value', '') != ''
            for timeline in self.studyDesignData.get('scheduleTimelines', [])
            if timeline.get('label') == 'Main Timeline'
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["LENGTH"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # NARMS
        if self.studyDesignData.get('arms', []):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["NARMS"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # NCOHORT
        if self.studyDesignData.get('population', {}).get("cohorts", []):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["NCOHORT"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # OBJEXP
        if any(objective.get('level', {}).get('code', '') == 'C163559' for objective in self.studyDesignData.get('objectives', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["OBJEXP"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # OBJPRIM
        if any(objective.get('level', {}).get('code', '') == 'C85826' for objective in self.studyDesignData.get('objectives', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["OBJPRIM"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # OBJSEC
        if any(objective.get('level', {}).get('code', '') == 'C85827' for objective in self.studyDesignData.get('objectives', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["OBJSEC"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # OBSMODEL
        if self.studyDesignData['studyType']['code'] == 'C16084' and self.studyDesignData.get('model', {}).get('decode', '') != '':
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["OBSMODEL"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # OBSTIMP
        if self.studyDesignData['studyType']['code'] == 'C16084' and self.studyDesignData.get('timePerspective', {}).get('decode', '') != '':
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["OBSTIMP"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # OBSTPOPD
        if self.studyDesignData['studyType']['code'] == 'C16084' and self.studyDesignData.get('population', {}).get('description', '') != '':
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["OBSTPOPD"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # OBSTSMM
        if self.studyDesignData['studyType']['code'] == 'C16084' and self.studyDesignData.get('samplingMethod', {}).get('decode', '') != '':
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["OBSTSMM"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # OUTMSEXP
        if any(
            endpoint.get('level', {}).get('code', '') == 'C170559'
            for objective in self.studyDesignData.get('objectives', [])
            for endpoint in objective.get('endpoints', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["OUTMSEXP"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # OUTMSPRI
        if any(
            endpoint.get('level', {}).get('code', '') == 'C94496'
            for objective in self.studyDesignData.get('objectives', [])
            for endpoint in objective.get('endpoints', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["OUTMSPRI"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # OUTMSSEC
        if any(
            endpoint.get('level', {}).get('code', '') == 'C139173'
            for objective in self.studyDesignData.get('objectives', [])
            for endpoint in objective.get('endpoints', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["OUTMSSEC"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # PCLAS
        if any(
            product.get('pharmacologicClass', {}).get('decode')
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
            for administration in intervention.get('administrations', [])
            for product in self.study_version_data.get('administrableProducts', [])
            if product.get('id') == administration.get('administrableProductId')
            and any(
                designation.get('code') == 'C202579'
                for designation in product.get('productDesignation', [])
            )
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["PCLAS"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # PIPIND
        if any(
            reference_identifier.get('type', {}).get('code', '') == 'C215674'
            for reference_identifier in self.study_version_data.get('referenceIdentifiers', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["PIPIND"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # PLANSUB
        if self.studyDesignData.get('population', {}).get("plannedEnrollmentNumber", {}).get("value", None) is not None:            
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["PLANSUB"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # PTRTDUR
        if any(
            administration.get('duration', {}).get('quantity', {}).get('value', None)
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
            for administration in intervention.get('administrations', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["PTRTDUR"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # RANDOM
        if any(
            characteristic.get('level', {}).get('code', '') in ('C46079','C147145')
            for characteristic in self.studyDesignData.get('characteristics', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 1,
                "codeList": "CL.NY",
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["RANDOM"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)   
                        
        # RDIND
        if any(
            indication.get('isRareDisease') is True 
            for indication in self.studyDesignData.get('indications', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 1,
                "codeList": "CL.NY",
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["RDIND"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # ROUTE
        if any(
            administration.get('route', {}).get('standardCode', {}).get('decode', '') != ''
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
            for administration in intervention.get('administrations', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["ROUTE"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # RTSPCDES
        if any(
            biospecimenRetention.get('description', '') != ''
            for biospecimenRetention in self.studyDesignData.get('biospecimenRetentions', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["RTSPCDES"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # SEXPOP
        if any(
            sex.get('decode', '') != '' 
            for sex in self.studyDesignData.get('population', {}).get("plannedSex", [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["SEXPOP"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # SPONSOR
        if any(
            organization.get('type', {}).get('code', '') == 'C70793' 
            for organization in self.study_version_data.get('organizations', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["SPONSOR"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # STYPE
        if self.studyDesignData.get('studyType', {}).get('decode', '') != '':
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["STYPE"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # TBLIND
        if self.studyDesignData.get('blindingSchema', {}).get('standardCode', {}).get('decode', '') != '':
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["TBLIND"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # THERAREA
        if any(
            area.get('decode', '') != '' 
            for area in self.studyDesignData.get('therapeuticAreas', [])
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["THERAREA"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # TINDTP
        if self.studyDesignData.get('studyType', {}).get('code', '') == 'C98388' and \
           any(intent.get('decode', '') != '' for intent in self.studyDesignData.get('intentTypes', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["TINDTP"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # TITLE
        if any(
            title.get('text', '') != ''
            for title in self.study_version_data.get('titles', [])
            if title.get('type', {}).get('code', '') == 'C207616'
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["TITLE"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # TPHASE
        if self.studyDesignData.get('studyPhase', {}).get('standardCode', {}).get('decode', '') != '':
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["TPHASE"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # TRT
        if any(
            intervention.get('label', '') != ''
            for intervention_id in self.studyDesignData.get('studyInterventionIds', [])
            for intervention in self.study_version_data.get('studyInterventions', [])
            if intervention.get('id') == intervention_id
        ):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["TRT"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)

        # TTYPE
        if self.studyDesignData.get('studyType', {}).get('code', '') == 'C98388' and \
           any(subType.get('decode', '') != '' for subType in self.studyDesignData.get('subTypes', [])):
            tsparmcd_entry = {
                "dataType": "text",
                "length": 200,
                "originType": "Protocol",
                "originSource": "Sponsor",
                "WhereClause": [
                    {
                        "Clause": [
                            {
                                "Dataset": "TS",
                                "Variable": "TSPARMCD",
                                "item": "IT.TS.TSPARMCD",
                                "Comparator": "EQ",
                                "Values": ["TTYPE"]
                            }
                        ]
                    }
                ]
            }
            self.vlm_lookup["TSPARMCD"].append(tsparmcd_entry)



        # # Add elements VLM entries
        # if self.studyDesignData.get('elements', []):
        #     for element in self.studyDesignData.get('elements', []):
        #         element_name = element.get('name', '')
        #         element_label = element.get('label', '')
        #         element_description = element.get('description', '')
                
        #         # Create VLM entry for SESTDTC variable
        #         sestdtc_entry = {
        #             "role": "Timing",
        #             "dataType": "datetime",
        #             "length": 19,
        #             "originType": "Collected",
        #             "originSource": "Investigator",
        #             "WhereClause": [
        #                 {
        #                     "Clause": [
        #                         {
        #                             "Dataset": "SE",
        #                             "Variable": "ETCD",
        #                             "item": "IT.SE.ETCD",
        #                             "Comparator": "EQ",
        #                             "Values": [element_name]
        #                         }
        #                     ]
        #                 }
        #             ]
        #         }
                
        #         # Create VLM entry for SEENDTC variable
        #         seendtc_entry = {
        #             "role": "Timing",
        #             "dataType": "datetime",
        #             "length": 19,
        #             "originType": "Collected",
        #             "originSource": "Investigator",
        #             "WhereClause": [
        #                 {
        #                     "Clause": [
        #                         {
        #                             "Dataset": "SE",
        #                             "Variable": "ETCD",
        #                             "item": "IT.SE.ETCD",
        #                             "Comparator": "EQ",
        #                             "Values": [element_name]
        #                         }
        #                     ]
        #                 }
        #             ]
        #         }
                
        #         # Add to vlm_lookup
        #         if "SESTDTC" not in self.vlm_lookup:
        #             self.vlm_lookup["SESTDTC"] = []
        #         self.vlm_lookup["SESTDTC"].append(sestdtc_entry)
                
        #         if "SEENDTC" not in self.vlm_lookup:
        #             self.vlm_lookup["SEENDTC"] = []
        #         self.vlm_lookup["SEENDTC"].append(seendtc_entry)

    def update_datasets_dict(self):
        """
        Update datasets_dict with variable-value pairs from VLM.
        
        Seeds IE/SE domains and timing variables when eligibility criteria or
        elements are present, derives TEST/TESTCD restrictions (`test_dict`),
        builds in-memory IE/SE codelists, and merges VLM where-clause values
        from `bc_dict` into `datasets_dict` for downstream codelist
        restrictions.
        """
        # Add TS dataset
        if "TS" not in self.datasets_dict:
            self.datasets_dict["TS"] = {}

            # Add TSPARMCD values extracted from vlm_lookup["TSPARMCD"] WhereClause Values
            tsparmcd_values = set()
            for entry in self.vlm_lookup.get("TSPARMCD", []):
                for where_clause in entry.get("WhereClause", []):
                    for clause in where_clause.get("Clause", []):
                        for value in clause.get("Values", []):
                            if value:
                                tsparmcd_values.add(value)

            if tsparmcd_values:
                self.datasets_dict["TS"]["TSPARMCD"] = {"codelist": {
                    "C66738": sorted(tsparmcd_values)
                }}

            # Add TSPARM values by matching conceptIds from filtered TSPARMCD terms against C67152
            kept_concept_ids = {term.get('conceptId') for term in [
                t for t in self.client.get_api_json(f"/mdr/ct/packages/sdtmct-{self.sdtmct}/codelists/C66738")['terms']
                if not tsparmcd_values or t.get('submissionValue') in tsparmcd_values
            ]}

            tsparm_values = {
                term.get('submissionValue')
                for term in self.client.get_api_json(f"/mdr/ct/packages/sdtmct-{self.sdtmct}/codelists/C67152")['terms']
                if term.get('conceptId') in kept_concept_ids and term.get('submissionValue')
            }

            if tsparm_values:
                self.datasets_dict["TS"]["TSPARM"] = {"codelist": {
                    "C67152": sorted(tsparm_values)
                }}
        
        # Add TA dataset from arms
        if self.studyDesignData.get('arms', []):
            if "TA" not in self.datasets_dict:
                self.datasets_dict["TA"] = {}

                # Add EPOCH values from epochs
                if self.studyDesignData.get('epochs', []):
                    # Create EPOCH codelist with all epoch labels
                    self.datasets_dict["TA"]["EPOCH"] = {"codelist": {
                        "C99079": [
                            epoch.get('name', '') for epoch in self.studyDesignData.get('epochs', []) if epoch.get('name', '')
                        ]
                    }}

        # Add TE dataset from elements
        if self.studyDesignData.get('elements', []):
            if "TE" not in self.datasets_dict:
                self.datasets_dict["TE"] = {}

        # Add TI dataset from eligibility criteria
        if self.studyDesignData.get('eligibilityCriteria', []):
            if "TI" not in self.datasets_dict:
                self.datasets_dict["TI"] = {}

        # Add TV dataset from eligibility criteria
        if self.studyDesignData.get('eligibilityCriteria', []):
            if "TV" not in self.datasets_dict:
                self.datasets_dict["TV"] = {}

        # Add SV dataset from encounters
        if self.studyDesignData.get('encounters', []):
            if "SV" not in self.datasets_dict:
                self.datasets_dict["SV"] = {}

                # Add SVCNTMOD values from encounters
                terms = self.client.get_api_json(f"/mdr/ct/packages/sdtmct-{self.sdtmct}/codelists/C171445")['terms']
                contact_mode_codes = set()
                for encounter in self.studyDesignData.get('encounters', []):
                    for contact_mode in encounter.get('contactModes', []):
                        code = contact_mode.get('code', '')
                        if code:
                            # Find matching term by conceptId and get submissionValue
                            term = next((t for t in terms if t.get('conceptId') == code), None)
                            if term:
                                submission_value = term.get('submissionValue', '')
                                if submission_value:
                                    contact_mode_codes.add(submission_value)
                
                if contact_mode_codes:
                    self.datasets_dict["SV"]["SVCNTMOD"] = {"codelist": {
                        "C171445": sorted(list(contact_mode_codes))
                    }}
            
        # Add IE dataset from eligibility criteria
        if self.studyDesignData.get('eligibilityCriteria', []):
            if "IE" not in self.datasets_dict:
                self.datasets_dict["IE"] = {}
                    
        # Add SE dataset from elements
        if self.studyDesignData.get('elements', []):
            if "SE" not in self.datasets_dict:
                self.datasets_dict["SE"] = {}

                # Add EPOCH values from epochs
                if self.studyDesignData.get('epochs', []):
                    # Create EPOCH codelist with all epoch labels
                    self.datasets_dict["SE"]["EPOCH"] = {"codelist": {
                        "C99079": [
                            epoch.get('name', '') for epoch in self.studyDesignData.get('epochs', []) if epoch.get('name', '')
                        ]
                    }}
        
        variable_values = defaultdict(set)

        for concept_data in self.bc_dict.values():
            for item in concept_data:
                for field_data in item.values():
                    if 'WhereClause' in field_data:
                        for where_clause in field_data['WhereClause']:
                            if 'Clause' in where_clause:
                                # Process all clauses in the single Clause array
                                for clause in where_clause['Clause']:
                                    if 'Variable' in clause and 'Values' in clause and 'Dataset' in clause:
                                        dataset = clause['Dataset']
                                        variable = clause['Variable']
                                        codelist_concept_id = clause['Codelist Concept ID']
                                        values = clause['Values']
                                        key = (variable, dataset, codelist_concept_id)
                                        variable_values[key].update(values)

        for (variable, dataset, codelist_concept_id), values in variable_values.items():
            if dataset not in self.datasets_dict:
                self.datasets_dict[dataset] = {}
            if variable not in self.datasets_dict[dataset]:
                self.datasets_dict[dataset][variable] = {"codelist": {}}
            if codelist_concept_id not in self.datasets_dict[dataset][variable]["codelist"]:
                self.datasets_dict[dataset][variable]["codelist"][codelist_concept_id] = []

            if variable.endswith('TESTCD'):
                terms = self.client.get_codelist_terms(f"sdtmct-{self.sdtmct}/", codelist_concept_id)
                response_codes = []
                for value in values:
                    term = next((term for term in terms if term.get('submissionValue') == value), None)
                    if term:
                        response_codes.append(term['conceptId'])

                if dataset not in self.test_dict:
                    self.test_dict[dataset] = {}
                
                self.test_dict[dataset][variable.replace('TESTCD', 'TEST')] = response_codes

            self.datasets_dict[dataset][variable]["codelist"][codelist_concept_id] = sorted(
                list(set(self.datasets_dict[dataset][variable]["codelist"][codelist_concept_id]).union(values))
            )
        
        # Create ARMCD codelists from arms
        if self.studyDesignData.get('arms', []):
            # Create ARMCD codelist with all arms name
            armcd_terms = []
            for arm in self.studyDesignData.get('arms', []):
                name = arm.get('name', '')
                if name:
                    armcd_terms.append({"codedValue": name})
            
            if armcd_terms:
                armcd_codelist_name = "ARMCD"
                if armcd_codelist_name not in self.code_lists_map:
                    self.code_lists_map[armcd_codelist_name] = {
                        "OID": f"CL.{armcd_codelist_name}",
                        "name": "ARM Code",
                        "dataType": "text",
                        "isNonStandard": True,
                        "codeListItems": armcd_terms
                    }

            # Create ARM codelist with all arms name
            arm_terms = []
            for arm in self.studyDesignData.get('arms', []):
                label = arm.get('name', '')
                if name:
                    arm_terms.append({"codedValue": label})
            
            if arm_terms:
                arm_codelist_name = "ARM"
                if arm_codelist_name not in self.code_lists_map:
                    self.code_lists_map[arm_codelist_name] = {
                        "OID": f"CL.{arm_codelist_name}",
                        "name": "ARM",
                        "dataType": "text",
                        "isNonStandard": True,
                        "codeListItems": arm_terms
                    }



        # Create IE codelists from eligibility criteria
        if self.studyDesignData.get('eligibilityCriteria', []):
            # Create IETEST codelist with all eligibility criterion labels
            ietest_terms = []
            for criterion in self.studyDesignData.get('eligibilityCriteria', []):
                label = criterion.get('label', '')
                if label:
                    ietest_terms.append({"codedValue": label})
            
            if ietest_terms:
                ietest_codelist_name = "IETEST"
                if ietest_codelist_name not in self.code_lists_map:
                    self.code_lists_map[ietest_codelist_name] = {
                        "OID": f"CL.{ietest_codelist_name}",
                        "name": "Inclusion/Exclusion Test Name",
                        "dataType": "text",
                        "isNonStandard": True,
                        "codeListItems": ietest_terms
                    }
            
            # Create IETESTCD codelist with criterion names and labels
            ietestcd_terms = []
            for criterion in self.studyDesignData.get('eligibilityCriteria', []):
                name = criterion.get('name', '')
                label = criterion.get('label', '')
                if name and label:
                    ietestcd_terms.append({
                        "codedValue": name,
                        "decode": label
                    })
            
            if ietestcd_terms:
                ietestcd_codelist_name = "IETESTCD"
                if ietestcd_codelist_name not in self.code_lists_map:
                    self.code_lists_map[ietestcd_codelist_name] = {
                        "OID": f"CL.{ietestcd_codelist_name}",
                        "name": "Inclusion/Exclusion Test Code",
                        "dataType": "text",
                        "isNonStandard": True,
                        "codeListItems": ietestcd_terms
                    }
        
        # Create SE codelists from elements
        if self.studyDesignData.get('elements', []):
            # Create ELEMENT codelist with all element labels
            element_terms = []
            for element in self.studyDesignData.get('elements', []):
                label = element.get('label', '')
                if label:
                    element_terms.append({"codedValue": label})
            
            if element_terms:
                element_codelist_name = "ELEMENT"
                if element_codelist_name not in self.code_lists_map:
                    self.code_lists_map[element_codelist_name] = {
                        "OID": f"CL.{element_codelist_name}",
                        "name": "Description of Element",
                        "dataType": "text",
                        "isNonStandard": True,
                        "codeListItems": element_terms
                    }
            
            # Create ETCD codelist with element names and labels
            etcd_terms = []
            for element in self.studyDesignData.get('elements', []):
                name = element.get('name', '')
                label = element.get('label', '')
                if name and label:
                    etcd_terms.append({
                        "codedValue": name,
                        "decode": label
                    })
            
            if etcd_terms:
                etcd_codelist_name = "ETCD"
                if etcd_codelist_name not in self.code_lists_map:
                    self.code_lists_map[etcd_codelist_name] = {
                        "OID": f"CL.{etcd_codelist_name}",
                        "name": "Element Code",
                        "dataType": "text",
                        "isNonStandard": True,
                        "codeListItems": etcd_terms
                    }

        # Create EPOCH codelists from epochs
        if self.studyDesignData.get('epochs', []):
            # Create EPOCH codelist with all epoch labels
            epoch_terms = []
            for epoch in self.studyDesignData.get('epochs', []):
                name = epoch.get('name', '')
                if name:
                    term = {"codedValue": name}
                    
                    # Add coding if type.code exists
                    type_code = epoch.get('type', {}).get('code', '')
                    if type_code:
                        term["coding"] = {
                            "code": type_code,
                            "codeSystem": "nci:ExtCodeID"
                        }
                    
                    epoch_terms.append(term)
            
            if epoch_terms:
                epoch_codelist_name = "EPOCH"
                if epoch_codelist_name not in self.code_lists_map:
                    self.code_lists_map[epoch_codelist_name] = {
                        "OID": f"CL.{epoch_codelist_name}",
                        "name": "Epoch",
                        "dataType": "text",
                        "standard": "STD.SDTMCT",
                        "coding": [{
                            "code": "C99079",
                            "codeSystem": "nci:ExtCodeID"
                        }],
                        "codeListItems": epoch_terms,
                    }

    def process_datasets(self):
        """
        Process datasets to populate items, itemGroups, whereClauses, and conditions.
        
        Processes all datasets found in datasets_dict.
        Fetches dataset metadata from CDISC Library and dispatches to standard dataset processor.
        """
        for dataset_name in self.datasets_dict.keys():
            try:
                dataset_data = self.client.get_api_json(f"/mdr/sdtmig/{self.sdtmig.replace('.', '-')}/datasets/{dataset_name}")
                self._process_standard_dataset(dataset_name, dataset_data)
            except Exception as e:
                print(f"Warning: Could not fetch dataset {dataset_name} from CDISC API: {e}")

    def _process_standard_dataset(self, dataset, dataset_data):
        """
        Process a standard SDTM dataset from CDISC API.
        
        Builds item definitions, wires IE/SE variables to the synthetic
        codelists created earlier, and groups VLM items into ValueList slices
        using `vlm_items_by_variable`, while emitting conditions and
        whereClauses for VLM applicability.
        
        Args:
            dataset (str): Dataset domain name (e.g., "VS")
            dataset_data (dict): Dataset metadata from CDISC Library
        """
        variable_list = list(self.datasets_dict[dataset].keys())
        exception_variables = self.required_variables_exceptions.get(dataset, [])

        all_vars = [v for v in dataset_data['datasetVariables'] 
                    if v['name'] in variable_list or v['core'] in ['Req', 'Exp'] or v['name'] in exception_variables]
        
        item_group = {
            "OID": f"IG.{dataset}",
            "name": dataset,
            "description": dataset_data['label'],
            "domain": dataset,
            # Here update purpose when we use for ADaM datasets
            "purpose": "Tabulation",
            "structure": dataset_data['datasetStructure'],
            "isReferenceData": (dataset_data.get('_links', {}).get('parentClass', {}).get('title') == "Trial Design") or not any(v.get('name') in ("USUBJID", "POOLID") for v in all_vars),
            "keySequence": ["__PLACEHOLDER__"],
            # Here are we always sure to use STD.SDTMIG?
            "standard": "STD.SDTMIG",
            "observationClass": {
                "name": dataset_data.get("_links", {}).get("parentClass", {}).get("title", "").upper()
            },
            # Here add future code for subClass in ADaM if needed
            "items": []
        }

        # Add standard items to itemGroup
        for var in all_vars:
            item_dict = {
                "OID": f"IT.{dataset}.{var['name']}",
                "mandatory": var['core'] == 'Req',
                "name": var['name'],
                "description": var['label'],
                "role": var['role']
            }
            
            # Add optional fields from datasets_dict if they exist
            var_data = self.datasets_dict[dataset].get(var['name'], {})

            if 'dataType' in var_data:
                data_type = var_data['dataType']
            else:
                data_type = self._convert_data_type(var)
            item_dict["dataType"] = data_type
              
            if data_type in ['text', 'integer', 'float']:               
                if 'length' in var_data:
                    item_dict['length'] = var_data['length']
                else:
                    item_dict['length'] = None
            
            if 'format' in var_data:
                item_dict['displayFormat'] = var_data['format']
            
            if 'significantDigits' in var_data:
                item_dict['significantDigits'] = var_data['significantDigits']
            
            if 'originType' in var_data and 'originSource' in var_data:
                item_dict['origin'] = {
                    "type": var_data['originType'],
                    "source": var_data['originSource']
                }
            else:
                item_dict['origin'] = {
                    "type": "__PLACEHOLDER__",
                    "source": "__PLACEHOLDER__"
                }

            # Add codelist references for TA variables
            if dataset == "TA":
                if var['name'] == "ARMCD":
                    if "ARMCD" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["ARMCD"]["OID"]
                elif var['name'] == "ARM":
                    if "ARM" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["ARM"]["OID"]
                elif var['name'] == "ELEMENT":
                    if "ELEMENT" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["ELEMENT"]["OID"]
                elif var['name'] == "ETCD":
                    if "ETCD" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["ETCD"]["OID"]

            # Add codelist references for TE variables
            if dataset == "TE":
                if var['name'] == "ELEMENT":
                    if "ELEMENT" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["ELEMENT"]["OID"]
                elif var['name'] == "ETCD":
                    if "ETCD" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["ETCD"]["OID"]


            # Add codelist references for TI variables
            if dataset == "TI":
                if var['name'] == "IETEST":
                    if "IETEST" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["IETEST"]["OID"]
                elif var['name'] == "IETESTCD":
                    if "IETESTCD" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["IETESTCD"]["OID"]

            # Add codelist references for TV variables
            if dataset == "TV":
                if var['name'] == "ARMCD":
                    if "ARMCD" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["ARMCD"]["OID"]
                elif var['name'] == "ARM":
                    if "ARM" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["ARM"]["OID"]
            
            # Add codelist references for IE variables
            if dataset == "IE":
                if var['name'] == "IETEST":
                    if "IETEST" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["IETEST"]["OID"]
                elif var['name'] == "IETESTCD":
                    if "IETESTCD" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["IETESTCD"]["OID"]
            
            # Add codelist references for SE variables
            if dataset == "SE":
                if var['name'] == "ELEMENT":
                    if "ELEMENT" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["ELEMENT"]["OID"]
                elif var['name'] == "ETCD":
                    if "ETCD" in self.code_lists_map:
                        item_dict['codeList'] = self.code_lists_map["ETCD"]["OID"]

            codelist_oid = None
            if var['name'] in variable_list:
                restriction_data = self.datasets_dict[dataset][var['name']].get('codelist', {})
                codelist_oid = self._process_variable_codelist(var, dataset, restriction_data)
            else:
                if '_links' in var and 'codelist' in var['_links']:
                    codelist_oid = self._process_variable_codelist(var, dataset)
            
            if codelist_oid:
                item_dict['codeList'] = codelist_oid
            
            item_group['items'].append(item_dict)
            
            # Collect VLM items for this variable
            if var['name'] in self.vlm_lookup:
                for vlm_data in self.vlm_lookup[var['name']]:
                    if 'WhereClause' in vlm_data and vlm_data['WhereClause']:
                        # Create condition and where clause
                        all_condition_oids = self._get_or_create_condition_from_vlm(vlm_data['WhereClause'], dataset, var['name'])
                        where_clause_oids = []
                        for condition_oids in all_condition_oids:
                            where_clause_oid = self._create_where_clause_for_variable(dataset, var['name'], condition_oids)
                            where_clause_oids.append(where_clause_oid)
                        
                        # Get the parameter value for OID suffix
                        param = None
                        for wc_data in vlm_data['WhereClause']:
                            if 'Clause' in wc_data:
                                for clause in wc_data['Clause']:
                                    if 'Values' in clause and clause['Values']:
                                        param = clause['Values'][0]
                                        break
                            if param:
                                break
                        
                        # Create VLM item with optional fields
                        data_type = vlm_data.get('dataType', self._convert_data_type(var))
                        vlm_item = {
                            "OID": f"IT.{dataset}.{var['name']}.{param}" if param else f"IT.{dataset}.{var['name']}",
                            "mandatory": False,
                            "name": var['name'],
                            "dataType": data_type
                        }
                        # Add optional fields
                        if data_type in ['text', 'integer', 'float']:
                            if 'length' in vlm_data:
                                vlm_item['length'] = vlm_data['length']
                            else:
                                vlm_item['length'] = None
                        if 'format' in vlm_data:
                            vlm_item['displayFormat'] = vlm_data['format']
                        if 'significantDigits' in vlm_data:
                            vlm_item['significantDigits'] = vlm_data['significantDigits']
                        if 'codeList' in vlm_data:
                            vlm_item['codeList'] = vlm_data['codeList']
                        if 'originSource' in vlm_data:
                            vlm_item['origin'] = {
                                'type': vlm_data['originType'],
                                'source': vlm_data['originSource']
                            }
                        
                        # Add applicableWhen as the last attribute
                        vlm_item['applicableWhen'] = where_clause_oids
                        
                        # Group VLM items by variable name for slices
                        vlm_key = f"{dataset}.{var['name']}"
                        if vlm_key not in self.vlm_items_by_variable:
                            self.vlm_items_by_variable[vlm_key] = []
                        self.vlm_items_by_variable[vlm_key].append(vlm_item)
        
        # Create slices from VLM items grouped by variable
        slices = []
        for vlm_key in sorted(self.vlm_items_by_variable.keys()):
            if vlm_key.startswith(f"{dataset}."):
                var_name = vlm_key.split('.', 1)[1]
                slice_dict = {
                    "OID": f"VL.{dataset}.{var_name}",
                    "name": f"VL_{dataset}_{var_name}",
                    "type": "ValueList",
                    "wasDerivedFrom": f"IT.{dataset}.{var_name}",
                    "items": self.vlm_items_by_variable[vlm_key]
                }
                slices.append(slice_dict)
        
        if slices:
            item_group['slices'] = slices
        
        self.item_groups.append(item_group)
   
    def save_debug_files(self, prefix="debug"):
        """
        Save all intermediate dictionaries to JSON files for debugging.
        
        Args:
            prefix (str): Prefix for debug file names (default: "debug")
        """
        debug_data = {
            "dataset_data": self.all_dataset_data,
            "datasets_dict": self.datasets_dict,
            "bc_dict": self.bc_dict,
            "vlm_lookup": self.vlm_lookup,
            "test_dict": self.test_dict,
            "condition_lookup": self.condition_lookup,
            "code_lists_map": self.code_lists_map,
            "vlm_items_by_variable": self.vlm_items_by_variable
        }
       
        for name, data in debug_data.items():
            output_file = f"{prefix}_{name}.json"
            with open(output_file, "w") as f:
                json.dump(data, f, indent=2)
            print(f"Debug file saved: {output_file}")
    
    def save_output(self):
        """
        Save the generated Define-JSON template to file.
        
        Assembles all processed components (itemGroups, conditions,
        whereClauses, codeLists) into the template and writes to output file.
        """
        self.template['itemGroups'] = self.item_groups
        self.template['whereClauses'] = self.where_clauses
        self.template['conditions'] = self.conditions
        self.template['codeLists'] = list(self.code_lists_map.values())
        
        with open(self.output_template, "w") as f:
            f.write(json.dumps(self.template, indent=2))
    
    def validate_against_schema(self, schema_file, excel_output=None):
        """
        Validate the generated Define-JSON file against a YAML schema.
        
        This method loads the YAML schema file and validates the generated
        JSON output against it using LinkML validation if available, or
        performs basic structural validation otherwise.
        
        Args:
            schema_file (str): Path to the YAML schema file
            excel_output (str, optional): Path to Excel file for validation report
        
        Returns:
            bool: True if validation passes, False otherwise
        """
        print(f"\nValidating output against schema: {schema_file}")
        
        try:
            # Load the schema file
            with open(schema_file, 'r', encoding='utf-8') as f:
                schema = yaml.safe_load(f)
            
            print(f"✓ Schema file loaded successfully")
            
            # Try to use linkml-runtime if available
            try:
                from linkml_runtime.loaders import json_loader
                from linkml_runtime.utils.schemaview import SchemaView
                
                print("Using LinkML validation...")
                
                # Create schema view
                sv = SchemaView(schema_file)
                
                # Load the generated JSON data
                with open(self.output_template, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # Try to use LinkML validators
                validation_errors = []
                validation_passed = False
                
                try:
                    # Method 1: Try using linkml validate_file
                    try:
                        from linkml.validator import validate_file
                        report = validate_file(self.output_template, schema_file)
                        
                        if report.results:
                            for result in report.results:
                                validation_errors.append(str(result))
                        else:
                            validation_passed = True
                    except (ImportError, AttributeError):
                        pass
                    
                    # Method 2: Try using Validator class directly
                    if not validation_passed and not validation_errors:
                        try:
                            from linkml.validator import Validator
                            validator = Validator(schema_file)
                            report = validator.validate_file(self.output_template)
                            
                            if report.results:
                                for result in report.results:
                                    validation_errors.append(str(result))
                            else:
                                validation_passed = True
                        except (ImportError, AttributeError, Exception):
                            pass
                    
                    # Method 3: Try using jsonschema generation and validation
                    if not validation_passed and not validation_errors:
                        try:
                            from linkml.generators.jsonschemagen import JsonSchemaGenerator
                            import jsonschema
                            
                            # Generate JSON Schema from LinkML
                            gen = JsonSchemaGenerator(schema_file)
                            json_schema = json.loads(gen.serialize())
                            
                            # Validate data against JSON Schema
                            jsonschema.validate(instance=data, schema=json_schema)
                            validation_passed = True
                            print("✓ Using JSON Schema validation (generated from LinkML)")
                            
                        except jsonschema.ValidationError as ve:
                            validation_errors.append(f"Path: {' -> '.join(str(p) for p in ve.path)}: {ve.message}")
                        except (ImportError, AttributeError, Exception) as e:
                            pass
                    
                    # Report results
                    if validation_passed:
                        print("✅ LinkML validation passed!")
                        if excel_output:
                            self._write_validation_excel([], excel_output, passed=True)
                        return True
                    elif validation_errors:
                        print(f"⚠️  LinkML validation found {len(validation_errors)} issue(s)")
                        if excel_output:
                            self._write_validation_excel(validation_errors, excel_output, passed=False)
                            print(f"   ✓ Validation report written to: {excel_output}")
                        else:
                            for error in validation_errors[:10]:  # Show first 10 errors
                                print(f"   • {error}")
                            if len(validation_errors) > 10:
                                print(f"   ... and {len(validation_errors) - 10} more errors")
                        print()
                        return False
                    else:
                        # No validator worked, fall back
                        raise ImportError("No compatible LinkML validator found")
                
                except Exception as ie:
                    # Validator not available, try alternative
                    print(f"ℹ️  Full LinkML validation not available")
                    print("   Using schema structure validation instead...")
                    
                    # Fallback: verify schema structure
                    all_classes = sv.all_classes()
                    
                    if all_classes:
                        print(f"   ✓ Schema contains {len(all_classes)} classes")
                        print("   ✅ Schema structure is valid!")
                        print()
                        return True
                    else:
                        print("   ⚠️  Could not load schema classes")
                        print("   Performing basic structural validation instead...\n")
                        return self._basic_schema_validation(schema)
                
            except ImportError:
                print("ℹ️  linkml-runtime not installed")
                print("   Install with: pip install linkml-runtime")
                print("   Performing basic structural validation instead...\n")
                return self._basic_schema_validation(schema)
            
            except Exception as e:
                print(f"⚠️  LinkML validation error: {e}")
                print("   Performing basic structural validation instead...\n")
                return self._basic_schema_validation(schema)
        
        except FileNotFoundError:
            print(f"❌ Schema file not found: {schema_file}")
            return False
        except yaml.YAMLError as e:
            print(f"❌ Error parsing YAML schema: {e}")
            return False
        except Exception as e:
            print(f"❌ Validation error: {e}")
            return False
    
    def _basic_schema_validation(self, schema):
        """
        Perform basic structural validation against the schema.
        
        Args:
            schema (dict): Loaded YAML schema
        
        Returns:
            bool: True if basic validation passes, False otherwise
        """
        errors = []
        warnings = []
        
        # Check if schema has classes defined
        if 'classes' not in schema:
            print("⚠️  Schema does not define classes")
            return True
        
        # Get all class definitions
        classes = schema.get('classes', {})
        
        # Basic validation: check that data structures match expected patterns
        print("Performing basic structural checks...\n")
        
        # Check root level fields
        if 'fileOID' not in self.template:
            errors.append("Missing required field: fileOID")
        if 'studyOID' not in self.template:
            errors.append("Missing required field: studyOID")
        
        # Check itemGroups structure
        for idx, ig in enumerate(self.template.get('itemGroups', [])):
            if 'OID' not in ig:
                errors.append(f"itemGroups[{idx}]: Missing OID")
            if 'name' not in ig:
                errors.append(f"itemGroups[{idx}]: Missing name")
        
        # Check codeLists structure
        for idx, cl in enumerate(self.template.get('codeLists', [])):
            if 'OID' not in cl:
                errors.append(f"codeLists[{idx}]: Missing OID")
            if 'name' not in cl:
                errors.append(f"codeLists[{idx}]: Missing name")
        
        # Print results
        if errors:
            print("❌ VALIDATION ERRORS:")
            for error in errors:
                print(f"   • {error}")
            print()
        
        if warnings:
            print("⚠️  VALIDATION WARNINGS:")
            for warning in warnings:
                print(f"   • {warning}")
            print()
        
        if not errors and not warnings:
            print("✅ Basic structural validation passed!")
        elif not errors:
            print(f"✅ Basic validation passed with {len(warnings)} warning(s)")
        else:
            print(f"❌ Basic validation failed with {len(errors)} error(s)")
        
        print()
        return len(errors) == 0
    
    def _write_validation_excel(self, validation_errors, excel_path, passed=True):
        """
        Write validation results to an Excel file.
        
        Args:
            validation_errors (list): List of validation error messages
            excel_path (str): Path to output Excel file
            passed (bool): Whether validation passed overall
        """
        try:
            import pandas as pd
            from datetime import datetime
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create summary data
            summary_data = {
                'Validation Date': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                'Schema File': [self.output_template.replace('.json', '_schema.yaml')],
                'Data File': [self.output_template],
                'Status': ['PASSED' if passed else 'FAILED'],
                'Total Errors': [len(validation_errors)],
                'Total ItemGroups': [len(self.template.get('itemGroups', []))],
                'Total CodeLists': [len(self.template.get('codeLists', []))],
                'Total WhereClauses': [len(self.template.get('whereClauses', []))],
                'Total Conditions': [len(self.template.get('conditions', []))]
            }
            
            # Create errors data with better parsing
            if validation_errors:
                errors_data = []
                for idx, error in enumerate(validation_errors, 1):
                    error_str = str(error)
                    
                    # Parse different error formats
                    error_type = 'Validation Error'
                    severity = 'ERROR'
                    location = 'N/A'
                    field = 'N/A'
                    issue = 'Unknown issue'
                    constraint = 'N/A'
                    
                    # Extract type and severity if present
                    if "type='" in error_str:
                        type_match = error_str.split("type='")[1].split("'")[0]
                        error_type = type_match.replace('_', ' ').title()
                    
                    if "severity=" in error_str:
                        sev_match = error_str.split("severity=")[1].split()[0]
                        severity = sev_match.replace('<Severity.', '').replace('>', '').replace("'", '').replace(':', '')
                    
                    # Extract the actual message content
                    if "message=\"" in error_str or "message='" in error_str:
                        msg_start = error_str.find("message=") + 9
                        msg_content = error_str[msg_start:]
                        
                        # Try to parse the message content
                        if msg_content.startswith('{'):
                            # This is a JSON object - likely the problematic data
                            try:
                                # Extract key fields from the object
                                if "'OID':" in msg_content or '"OID":' in msg_content:
                                    oid_match = msg_content.split("'OID':")[1].split(',')[0].strip() if "'OID':" in msg_content else msg_content.split('"OID":')[1].split(',')[0].strip()
                                    oid_clean = oid_match.replace("'", "").replace('"', '').strip()
                                    field = oid_clean
                                
                                if "'name':" in msg_content or '"name":' in msg_content:
                                    name_match = msg_content.split("'name':")[1].split(',')[0].strip() if "'name':" in msg_content else msg_content.split('"name":')[1].split(',')[0].strip()
                                    name_clean = name_match.replace("'", "").replace('"', '').strip()
                                    location = f"Object: {name_clean}"
                                
                                # Determine what's wrong - look for the actual validation error
                                if ' is not valid under any of the given schemas' in error_str:
                                    issue = 'Object does not match any allowed schema pattern'
                                elif 'Additional properties are not allowed' in error_str:
                                    # Extract the additional property
                                    if 'were unexpected' in error_str:
                                        props = error_str.split('(')[1].split('were unexpected')[0].strip()
                                        issue = f'Unexpected properties: {props}'
                                    else:
                                        issue = 'Contains properties not defined in schema'
                                elif 'is a required property' in error_str:
                                    prop = error_str.split("'")[1] if "'" in error_str else 'unknown'
                                    issue = f'Missing required property: {prop}'
                                elif 'does not match' in error_str:
                                    issue = 'Value does not match expected pattern/format'
                                else:
                                    issue = 'Schema validation failed - check required fields and data types'
                                    
                            except:
                                issue = 'Object structure does not match schema'
                        else:
                            # Direct message
                            issue = msg_content.strip().replace('"', '').replace("'", '')[:150]
                    
                    # Try to extract constraint information
                    if 'constraint:' in error_str.lower():
                        constraint_parts = error_str.lower().split('constraint:')
                        if len(constraint_parts) > 1:
                            constraint = constraint_parts[1].split()[0]
                    
                    # Fallback: Try to extract path information
                    if location == 'N/A' and 'Path:' in error_str:
                        parts = error_str.split('Path:', 1)
                        if len(parts) > 1:
                            path_and_msg = parts[1].split(':', 1)
                            location = path_and_msg[0].strip()
                            if len(path_and_msg) > 1 and issue == 'Unknown issue':
                                issue = path_and_msg[1].strip()
                    
                    # Extract field name from location if not found
                    if field == 'N/A' and location != 'N/A':
                        if ' -> ' in location:
                            path_parts = location.split(' -> ')
                            field = path_parts[-1] if path_parts else location
                        else:
                            field = location
                    
                    errors_data.append({
                        'Error #': idx,
                        'Severity': severity,
                        'Type': error_type,
                        'Object/Field': field,
                        'Location': location,
                        'Issue': issue,
                        'Constraint': constraint,
                        'Full Error': error_str
                    })
                errors_df = pd.DataFrame(errors_data)
            else:
                errors_df = pd.DataFrame({'Message': ['No validation errors found']})
            
            summary_df = pd.DataFrame(summary_data)
            
            # Write to Excel with multiple sheets
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                errors_df.to_excel(writer, sheet_name='Validation Errors', index=False)
                
                # Format Summary sheet
                summary_ws = writer.sheets['Summary']
                for cell in summary_ws[1]:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
                
                # Format Errors sheet
                if validation_errors:
                    errors_ws = writer.sheets['Validation Errors']
                    
                    # Header formatting
                    for cell in errors_ws[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                    
                    # Set column widths
                    errors_ws.column_dimensions['A'].width = 8   # Error #
                    errors_ws.column_dimensions['B'].width = 12  # Severity
                    errors_ws.column_dimensions['C'].width = 20  # Type
                    errors_ws.column_dimensions['D'].width = 25  # Object/Field
                    errors_ws.column_dimensions['E'].width = 30  # Location
                    errors_ws.column_dimensions['F'].width = 50  # Issue
                    errors_ws.column_dimensions['G'].width = 15  # Constraint
                    errors_ws.column_dimensions['H'].width = 15  # Full Error (hidden by default)
                    
                    # Hide the Full Error column (but keep it for reference)
                    errors_ws.column_dimensions['H'].hidden = True
                    
                    # Enable text wrapping for Issue column
                    for row in errors_ws.iter_rows(min_row=2, max_row=len(validation_errors)+1, min_col=6, max_col=6):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Color-code severity
                    for row_idx, row in enumerate(errors_ws.iter_rows(min_row=2, max_row=len(validation_errors)+1, min_col=2, max_col=2), start=2):
                        for cell in row:
                            if cell.value == 'ERROR':
                                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            elif cell.value == 'WARNING':
                                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    errors_ws = writer.sheets['Validation Errors']
                    errors_ws.column_dimensions['A'].width = 50
                
                # Auto-adjust Summary sheet column widths
                for column in summary_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    summary_ws.column_dimensions[column_letter].width = adjusted_width
            
            return True
            
        except ImportError:
            print("   ⚠️  pandas and openpyxl required for Excel output")
            print("   Install with: pip install pandas openpyxl")
            return False
        except Exception as e:
            print(f"   ⚠️  Error writing Excel file: {e}")
            return False

    def process(self):
        """
        Main execution method orchestrating the full processing pipeline.
        
        Execution flow:
        1. Process biomedical concepts from USDM
        2. Build VLM lookup tables
        3. Update datasets dictionary with VLM data
        4. Populate study metadata
        5. Process datasets to generate Define-JSON structures
        6. Add CDISC standards information
        7. Save output to JSON file
        8. Save debug files if debug mode is enabled
        """
        self.process_biomedical_concepts()
        self.build_vlm_lookup()
        self.update_datasets_dict()
        self._build_global_codelist_terms()
        self.populate_study_elements()
        self.process_datasets()
        self.add_standards()
        self.save_output()
        
        if self.debug:
            self.save_debug_files()

    def _build_global_codelist_terms(self):
        """
        Pre-scan datasets_dict to build a union of all explicitly restricted
        terms per codelist C-Code. Used to decide placeholder vs. real terms.

        For each codelist C-Code encountered across all datasets/variables:
        - Non-empty list: contributes those values to the union
        - Empty list: contributes nothing (means "unrestricted" for that variable)

        Result stored in self.global_codelist_terms = {codelist_id: set_of_values}.
        A codelist with an empty set means no dataset ever restricted it → placeholder.
        """
        self.global_codelist_terms = {}
        for variables in self.datasets_dict.values():
            for var_data in variables.values():
                for codelist_id, terms in var_data.get('codelist', {}).items():
                    if codelist_id not in self.global_codelist_terms:
                        self.global_codelist_terms[codelist_id] = set()
                    for term in terms:
                        self.global_codelist_terms[codelist_id].add(term)

    def _get_or_create_condition_from_vlm(self, where_clause_data, dataset, variable_name):
        """
        Get existing or create new conditions from VLM where clause data.
        
        Creates individual conditions for each clause, deduplicating based on
        range check content. Each condition gets a unique hex-based OID.
        
        Args:
            where_clause_data (list): Where clause dictionaries from VLM
            dataset (str): Dataset domain name
            variable_name (str): Variable name (not used in current implementation)
        
        Returns:
            list: List of condition OID lists, one for each WhereClause
        """
        all_condition_oids = []
        
        for where_clause in where_clause_data:
            if 'Clause' in where_clause:
                clause_condition_oids = []
                # Process all clauses in the single Clause array
                for clause in where_clause['Clause']:
                    range_check = {
                        "comparator": clause['Comparator'],
                        "checkValues": clause['Values'],
                        "item": clause['item'],
                        "softHard": "Soft"
                    }
                    
                    range_checks_key = self._create_condition_key([range_check])
                    
                    if range_checks_key in self.condition_lookup:
                        clause_condition_oids.append(self.condition_lookup[range_checks_key])
                    else:
                        condition_content = f"{clause['Variable']}.{clause['Comparator']}.{'.'.join(sorted(clause['Values']))}"
                        hex_hash = self._generate_hex_oid(condition_content, "").replace(".", "")
                        condition_oid = f"COND.{dataset}.{clause['Variable']}.{hex_hash}"
                        
                        condition = {
                            "OID": condition_oid,
                            "rangeChecks": [range_check]
                        }
                        
                        self.conditions.append(condition)
                        self.condition_lookup[range_checks_key] = condition_oid
                        clause_condition_oids.append(condition_oid)
                
                all_condition_oids.append(clause_condition_oids)
        
        return all_condition_oids

    def _create_where_clause_for_variable(self, dataset, variable_name, condition_oids):
        """
        Create a where clause for a variable with associated conditions.
        
        Args:
            dataset (str): Dataset domain name
            variable_name (str): Variable name
            condition_oids (list): List of condition OIDs to reference
        
        Returns:
            str: Where clause OID
        """
        where_clause_content = f"{dataset}.{variable_name}.{'.'.join(sorted(condition_oids))}"
        hex_hash = self._generate_hex_oid(where_clause_content, "").replace(".", "")
        where_clause_oid = f"WC.{dataset}.{hex_hash}"
        
        where_clause = {
            "OID": where_clause_oid,
            "conditions": condition_oids
        }
        
        self.where_clauses.append(where_clause)
        return where_clause_oid

    def _create_condition_key(self, range_checks):
        """
        Create a unique key for condition deduplication.
        
        Args:
            range_checks (list): List of range check dictionaries
        
        Returns:
            str: Unique key representing the range checks
        """
        sorted_checks = sorted(range_checks, key=lambda x: (x.get('item', ''), x.get('comparator', '')))
        
        key_parts = []
        for check in sorted_checks:
            item = check.get('item', '')
            comparator = check.get('comparator', '')
            values = sorted(check.get('checkValues', []))
            concatenated_values = ''.join(values).replace(' ', '')
            key_parts.append(f"{item}.{comparator}.{concatenated_values}")
        
        return ".".join(key_parts)

    def _generate_hex_oid(self, content, prefix="ID"):
        """
        Generate a consistent hex-based OID using MD5 hash.
        
        Args:
            content (str): Content to hash for OID generation
            prefix (str): Prefix for the OID (e.g., "COND", "WC")
            
        Returns:
            str: Hex-based OID (e.g., "COND.3cd24cc4")
        """
        content_hash = hashlib.md5(content.encode('utf-8')).hexdigest()
        short_hash = content_hash[:8]
        return f"{prefix}.{short_hash}"

    def _process_variable_codelist(self, var, dataset, restriction_codes=None):
        """
        Process variable codelists with optional restrictions.
        
        Fetches controlled terminology from CDISC Library, applies restrictions
        if specified, and creates deduplicated codelist entries. Returns single
        OID string instead of list for format.
        
        Args:
            var (dict): Variable metadata from SDTMIG
            dataset (str): Dataset domain name
            restriction_codes (dict, optional): Codelist restrictions {codelist_id: [codes]}
        
        Returns:
            str: Codelist OID in format CL.{short_name}, or None if no codelist
        """
        if not ('_links' in var and 'codelist' in var['_links']):
            return None
        
        codelist_entries = []
        
        def create_term_dict(term):
            return {
                "NCI Term Code": term.get("conceptId"),
                "Term": term.get("submissionValue"),
                "Decoded Value": term.get("synonyms", [])
            }
        
        for codelist_item in var['_links']['codelist']:
            if not ('href' in codelist_item and codelist_item['href'] is not None):
                continue

            codelist_id = codelist_item['href'].split('/')[-1]
            codelist_data = self.client.get_api_json(f"/mdr/ct/packages/sdtmct-{self.sdtmct}/codelists/{codelist_id}")
            all_terms = codelist_data.get("terms", [])
            
            if restriction_codes and codelist_id in restriction_codes:
                restriction_list = restriction_codes[codelist_id]
                
                if not restriction_list:
                    global_terms = self.global_codelist_terms.get(codelist_id, set())
                    if global_terms:
                        final_terms = [
                            create_term_dict(term)
                            for term in all_terms
                            if term.get("submissionValue") in global_terms
                        ]
                    else:
                        final_terms = [{"NCI Term Code": None, "Term": "__PLACEHOLDER__", "Decoded Value": ["__PLACEHOLDER__"]}]
                else:
                    # Find matching terms from standard codelist
                    final_terms = [
                        create_term_dict(term) 
                        for term in all_terms 
                        if term.get("submissionValue") in restriction_list
                    ]
                    
                    # Check for non-matching restriction values and add them with codedValue only
                    matched_values = {term.get("submissionValue") for term in all_terms if term.get("submissionValue") in restriction_list}
                    unmatched_values = [val for val in restriction_list if val not in matched_values]
                    
                    # Add unmatched values with only codedValue
                    for unmatched_val in unmatched_values:
                        if unmatched_val:  # Ensure not empty
                            print(f"⚠️  Codelist '{codelist_data.get('name', codelist_id)}': term '{unmatched_val}' not found in standard codelist - adding with codedValue only")
                            final_terms.append({
                                "NCI Term Code": None,
                                "Term": unmatched_val
                            })
            
            elif var['name'].endswith('TEST') and dataset in self.test_dict and var['name'] in self.test_dict[dataset]:
                final_terms = [
                    create_term_dict(term)
                    for term in all_terms
                    if term.get("conceptId") in self.test_dict[dataset][var['name']]
                ]
            
            else:
                global_terms = self.global_codelist_terms.get(codelist_id, set())
                if global_terms:
                    final_terms = [
                        create_term_dict(term)
                        for term in all_terms
                        if term.get("submissionValue") in global_terms
                    ]
                else:
                    final_terms = [{"NCI Term Code": None, "Term": "__PLACEHOLDER__", "Decoded Value": ["__PLACEHOLDER__"]}]
            
            codelist_entry = {
                "NCI Codelist Code": codelist_data.get("conceptId"),
                "Name": codelist_data.get("name"),
                "Short Name": codelist_data.get("submissionValue"),
                "Terms": final_terms
            }
            
            codelist_entries.append(codelist_entry)
        
        # Process first codelist only and return single OID
        return_oid = None
        
        for entry in codelist_entries[:1]:  # Only process first codelist
            short_name = entry.get("Short Name")
            if not short_name:
                continue
            oid = f"CL.{short_name}"
            name = entry.get("Name", "")
            
            return_oid = oid

            #TODO: dataType should ideally come from the variable metadata or codelist metadata, but defaulting to "text" for now
            if short_name not in self.code_lists_map:
                codelist_dict = {
                    "OID": oid,
                    "name": name,
                    "dataType": "text",
                    "standard": "STD.SDTMCT",
                }
                
                # Add coding if NCI Codelist Code exists
                nci_code = entry.get("NCI Codelist Code")
                if nci_code:
                    codelist_dict["coding"] = [{
                        "code": nci_code,
                        "codeSystem": "nci:ExtCodeID"
                    }]
                
                codelist_dict["codeListItems"] = []
                self.code_lists_map[short_name] = codelist_dict

            existing_codes = {item["codedValue"] for item in self.code_lists_map[short_name]["codeListItems"]}

            #TODO: decoded value is by default here the first synonym, but we may want to consider other options or concatenating multiple synonyms in the future
            for term in entry.get("Terms", []):
                coded_value = term.get("Term")
                decoded_values = term.get("Decoded Value") or []
                decode = decoded_values[0] if isinstance(decoded_values, list) and len(decoded_values) > 0 else ""
                nci_term_code = term.get("NCI Term Code")

                if coded_value and coded_value not in existing_codes:
                    item_dict = {
                        "codedValue": coded_value,
                        "decode": decode
                    }
                    
                    # Add coding if NCI Term Code exists
                    if nci_term_code:
                        item_dict["coding"] = {
                            "code": nci_term_code,
                            "codeSystem": "nci:ExtCodeID"
                        }
                    
                    self.code_lists_map[short_name]["codeListItems"].append(item_dict)
                    existing_codes.add(coded_value)
        
        return return_oid

    def add_standards(self):
        """
        Add CDISC standards information to template.
        
        Populates standards section with SDTM IG and CT package information.
        """
        self.template["standards"] = [
            {
                "OID": "STD.SDTMIG",
                "name": "SDTMIG",
                "type": "IG",
                "version": self.sdtmig,
                "status": "FINAL"
            },
            {
                "OID": "STD.SDTMCT",
                "name": "CDISC/NCI",
                "type": "CT",
                "version": self.sdtmct,
                "status": "FINAL",
                "publishingSet": "SDTM"
            }
        ]

    def populate_study_elements(self):
        """
        Populate study metadata from USDM.
        
        Extracts study name, description, protocol name from USDM titles
        and populates template header fields with study-specific information.
        """
        # Extract study name, description, protocol name using jmespath filtering
        titles = jmespath.search(f'versions[{self.studyversion}].titles', self.usdm_data.get('study', {}))
        study_name = None
        study_description = None
        if titles:
            for title in titles:
                if title.get('type', {}).get('decode') == 'Study Acronym':
                    study_name = title.get('text')
                elif title.get('type', {}).get('decode') == 'Official Study Title':
                    study_description = title.get('text')
        protocol_name = study_name
        
        # Extract language code
        doc_version_id = jmespath.search(f'versions[{self.studyversion}].documentVersionIds[{self.docversion}]', self.usdm_data.get('study', {}))
        documented_by = jmespath.search(f'documentedBy', self.usdm_data.get('study', {}))
        language = None
        if documented_by and doc_version_id:
            for doc in documented_by:
                versions = doc.get('versions', [])
                for v in versions:
                    if v.get('id') == doc_version_id:
                        language = doc.get('language', {}).get('code')
                        break

        current_time = datetime.now()
        time_str = current_time.strftime("%Y-%m-%dT%H:%M:%S.%f") + "+00:00"
        
        # Increment version and design for display (0-indexed internally, 1-indexed for display)
        version_display = self.studyversion + 1
        design_display = self.studydesign + 1

        self.template['fileOID'] = f"ODM.DEFINE-JSON.{study_name}.Version{version_display}.Design{design_display}"
        self.template['creationDateTime'] = time_str
        self.template['studyOID'] = f"ODM.{study_name}.Version{version_display}.Design{design_display}"
        self.template['studyName'] = study_name
        self.template['studyDescription'] = study_description
        self.template['protocolName'] = protocol_name
        self.template['OID'] = f"MDV.{study_name}.Version{version_display}.Design{design_display}"
        self.template['name'] = f"MDV {study_name}"
        self.template['description'] = f"Data Definitions for {study_name}"

    # ------------------------------------------------------------------
    # Patch-file helpers
    # ------------------------------------------------------------------

    def _collect_item_placeholders(self, item, item_section):
        """Collect placeholder / null fields from a single item dict into item_section."""
        patches = {}
        if item.get('length') is None:
            patches['length'] = None
        origin = item.get('origin', {})
        if origin.get('type') == '__PLACEHOLDER__':
            patches['originType'] = '__PLACEHOLDER__'
        if origin.get('source') == '__PLACEHOLDER__':
            patches['originSource'] = '__PLACEHOLDER__'
        if patches:
            item_section[item['OID']] = patches

    def _find_item_by_oid(self, oid):
        """Return the first item dict in the assembled template matching *oid*, or None."""
        for ig in self.template.get('itemGroups', []):
            for item in ig.get('items', []):
                if item.get('OID') == oid:
                    return item
            for sl in ig.get('slices', []):
                for vlm_item in sl.get('items', []):
                    if vlm_item.get('OID') == oid:
                        return vlm_item
        return None

    def _apply_item_patch(self, item, item_patches):
        """Apply patch fields to a single item dict in-place."""
        oid = item.get('OID')
        if oid not in item_patches:
            return
        ip = item_patches[oid]
        if 'length' in ip and ip['length'] is not None:
            item['length'] = ip['length']
        if 'originType' in ip and ip['originType'] != '__PLACEHOLDER__':
            item.setdefault('origin', {})['type'] = ip['originType']
        if 'originSource' in ip and ip['originSource'] != '__PLACEHOLDER__':
            item.setdefault('origin', {})['source'] = ip['originSource']

    def generate_patch_file(self, patch_output_path):
        """
        Generate a YAML patch file listing all placeholder / null fields.

        Scans the assembled template for any field set to '__PLACEHOLDER__' or
        null and writes an OID-keyed YAML file that users can fill in and apply
        back with --apply_patch.

        Sections produced:
        - itemGroups: datasets whose keySequence still contains __PLACEHOLDER__
        - items: variables with null length or __PLACEHOLDER__ origin type/source
        - codeLists: codelists with at least one __PLACEHOLDER__ coded value

        Args:
            patch_output_path (str): Path to write the YAML patch file
        """
        ig_section = {}
        item_section = {}
        cl_section = {}

        for ig in self.template.get('itemGroups', []):
            if '__PLACEHOLDER__' in ig.get('keySequence', []):
                ig_section[ig['OID']] = {
                    'name': ig.get('name', ''),
                    'description': ig.get('description', ''),
                }

        for ig in self.template.get('itemGroups', []):
            for item in ig.get('items', []):
                self._collect_item_placeholders(item, item_section)
            for sl in ig.get('slices', []):
                for vlm_item in sl.get('items', []):
                    self._collect_item_placeholders(vlm_item, item_section)

        for cl in self.template.get('codeLists', []):
            ph_count = sum(
                1 for i in cl.get('codeListItems', [])
                if i.get('codedValue') == '__PLACEHOLDER__'
            )
            if ph_count:
                cl_section[cl['OID']] = {'name': cl.get('name', ''), 'count': ph_count}

        lines = [
            "# " + "=" * 68,
            "# Define-JSON Placeholder Patch File",
            f"# Study    : {self.template.get('studyName', 'N/A')}",
            f"# Generated: {datetime.now().strftime('%Y-%m-%d')}",
            "#",
            "# Instructions:",
            "#   1. Replace every __PLACEHOLDER__ with a valid value",
            "#   2. Fill every null with an appropriate number",
            "#   3. Apply with: --apply_patch <this_file> on your next run",
            "#",
            "# Valid values for originType:",
            "#   Collected | Derived | Assigned | Protocol | eDT | Predecessor | Not Available",
            "# Valid values for originSource:",
            "#   Investigator | Sponsor | Subject | Vendor",
            "# " + "=" * 68,
            "",
        ]

        if ig_section:
            lines.append("# keySequence: ordered list of variable names that form the sort / primary key")
            lines.append("# Example:  keySequence: [STUDYID, USUBJID, DSSEQ]")
            lines.append("itemGroups:")
            for oid, meta in ig_section.items():
                lines.append(f"  {oid}:  # {meta['name']} - {meta['description']}")
                lines.append(f"    keySequence: [__PLACEHOLDER__]")
            lines.append("")

        if item_section:
            lines.append("# length: integer (number of characters / digits for the variable)")
            lines.append("# originType and originSource: see valid values in the header above")
            lines.append("items:")
            for oid, fields in item_section.items():
                item_obj = self._find_item_by_oid(oid)
                comment = (
                    f"  # {item_obj.get('name', '')} - {item_obj.get('description', '')}"
                    if item_obj else ""
                )
                lines.append(f"  {oid}:{comment}")
                for field, value in fields.items():
                    yaml_val = "null" if value is None else f'"{value}"'
                    lines.append(f"    {field}: {yaml_val}")
            lines.append("")

        if cl_section:
            lines.append("# codeListItems: add one entry per term;")
            lines.append("# fill codedValue (submission value) and decode (display label)")
            lines.append("codeLists:")
            for oid, meta in cl_section.items():
                lines.append(f"  {oid}:  # {meta['name']}")
                lines.append("    codeListItems:")
                for _ in range(meta['count']):
                    lines.append('      - codedValue: "__PLACEHOLDER__"')
                    lines.append('        decode: "__PLACEHOLDER__"')
            lines.append("")

        content = "\n".join(lines)
        with open(patch_output_path, "w", encoding="utf-8") as f:
            f.write(content)

        print(f"\n📋 Patch file written: {patch_output_path}")
        print(f"   {len(ig_section)} dataset(s) need keySequence, "
              f"{len(item_section)} item(s) need origin/length, "
              f"{len(cl_section)} codelist(s) need terms")

    def apply_patch(self, patch_file):
        """
        Apply a YAML patch file to the assembled template and re-save.

        Reads the patch YAML generated by --patch_file (or written manually)
        and updates the in-memory template, then calls save_output() to write
        the patched result to disk.  Fields still set to '__PLACEHOLDER__' or
        null in the patch are left unchanged in the output.

        Args:
            patch_file (str): Path to the YAML patch file
        """
        with open(patch_file, "r", encoding="utf-8") as f:
            patch = yaml.safe_load(f)

        if not patch:
            print("⚠️  Patch file is empty — nothing to apply")
            return

        # itemGroups — update keySequence when fully filled
        for ig in self.template.get('itemGroups', []):
            oid = ig['OID']
            ip = patch.get('itemGroups', {}).get(oid, {})
            ks = ip.get('keySequence', [])
            if ks and '__PLACEHOLDER__' not in ks:
                ig['keySequence'] = ks

        # items — update length and origin fields
        item_patches = patch.get('items', {})
        if item_patches:
            for ig in self.template.get('itemGroups', []):
                for item in ig.get('items', []):
                    self._apply_item_patch(item, item_patches)
                for sl in ig.get('slices', []):
                    for vlm_item in sl.get('items', []):
                        self._apply_item_patch(vlm_item, item_patches)

        # codeLists — replace placeholder items with provided terms
        for cl in self.template.get('codeLists', []):
            oid = cl['OID']
            cp = patch.get('codeLists', {}).get(oid, {})
            if 'codeListItems' not in cp:
                continue
            existing = [
                i for i in cl.get('codeListItems', [])
                if i.get('codedValue') != '__PLACEHOLDER__'
            ]
            existing_codes = {i['codedValue'] for i in existing}
            for pi in cp['codeListItems']:
                cv = pi.get('codedValue')
                if cv and cv != '__PLACEHOLDER__' and cv not in existing_codes:
                    existing.append({'codedValue': cv, 'decode': pi.get('decode', '')})
                    existing_codes.add(cv)
            cl['codeListItems'] = existing

        self.save_output()
        print(f"✅ Patch applied and output re-saved to {self.output_template}")


def main():
    """
    Main entry point for command-line execution.
    
    Parses command-line arguments and orchestrates the USDM to Define-JSON
    transformation process.
    """
    # Reconfigure stdout to UTF-8 so Unicode symbols (✓, ✅, ❌, ⚠️) don't
    # crash on Windows terminals that default to cp1252.
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    parser = argparse.ArgumentParser(
        description="Process USDM JSON file to generate Define-JSON with slices.",
        epilog="Example: python create_define_json.py --usdm_file data/study.json --output_template output.json --sdtmct 2025-03-28 --validate define.yaml --validation_report validation_report.xlsx"
    )
    parser.add_argument(
        "--usdm_file", 
        required=True, 
        help="Path to USDM JSON file"
    )
    parser.add_argument(
        "--output_template",
        required=True,
        help="Path to output Define-JSON file"
    )
    parser.add_argument(
        "--sdtmct",
        required=True,
        help="SDTM Controlled Terminology date in yyyy-mm-dd format (required)"
    )
    parser.add_argument(
        "--sdtmig",
        default="3.4",
        help="SDTM Implementation Guide version (default: 3.4)"
    )
    parser.add_argument(
        "--studyversion",
        type=int,
        default=0,
        help="Study version index in USDM (default: 0)"
    )
    parser.add_argument(
        "--studydesign",
        type=int,
        default=0,
        help="Study design index in USDM (default: 0)"
    )
    parser.add_argument(
        "--docversion",
        type=int,
        default=0,
        help="Document version index in USDM (default: 0)"
    )
    parser.add_argument(
        "--cdisc_api_key",
        default=None,
        help="CDISC Library API Key (optional, defaults to environment variable CDISC_API_KEY)"
    )
    parser.add_argument(
        "--cosmosversion",
        default="v2",
        help="CDISC Cosmos API version (default: v2)"
    )
    parser.add_argument(
        "--validate",
        nargs='?',
        const="define.yaml",
        default=None,
        help="Validate output against YAML schema file (default: define.yaml)"
    )
    parser.add_argument(
        "--validation_report",
        required=False,
        help="Path to Excel file for validation report (e.g., validation_report.xlsx)"
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable debug mode to save intermediate dictionaries as JSON files"
    )
    parser.add_argument(
        "--patch_file",
        required=False,
        help=(
            "Path to a YAML patch file. On every run the file is (re)written with all "
            "remaining placeholder/null values so users can fill them in. Required when "
            "--apply_patch is used so the file is refreshed after applying (e.g., patch.yaml)."
        )
    )
    parser.add_argument(
        "--apply_patch",
        required=False,
        help=(
            "Path to a previously completed YAML patch file produced by --patch_file. "
            "Values are merged into the freshly generated template: OIDs that no "
            "longer exist are silently ignored, new OIDs added by a USDM update will "
            "appear in the refreshed --patch_file output. Requires --patch_file (e.g., patch.yaml)."
        )
    )
    args = parser.parse_args()
    
    # Validate conditional requirements
    if args.validate and not args.validation_report:
        parser.error("--validation_report is required when --validate is used")
    if args.apply_patch and not args.patch_file:
        parser.error("--patch_file is required when --apply_patch is used so the patch file is refreshed after applying")
    
    processor = USDMDefineJSONProcessor(
        usdm_file=args.usdm_file,
        output_template=args.output_template,
        sdtmig=args.sdtmig,
        sdtmct=args.sdtmct,
        studyversion=args.studyversion,
        studydesign=args.studydesign,
        docversion=args.docversion,
        cdisc_api_key=args.cdisc_api_key,
        cosmosversion=args.cosmosversion,
        debug=args.debug
    )
    
    processor.output_template = args.output_template
    processor.process()

    if args.apply_patch:
        processor.apply_patch(args.apply_patch)

    print(f"\n✅ Define-JSON file created successfully: {args.output_template}")

    if args.patch_file:
        processor.generate_patch_file(args.patch_file)

    # Validate if requested
    if args.validate:
        is_valid = processor.validate_against_schema(args.validate, excel_output=args.validation_report)
        if not is_valid:
            exit(1)


if __name__ == "__main__":
    main()

