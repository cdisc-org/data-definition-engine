"""
Dataset-JSON Generator from Define-XML
========================================

This module processes CDISC Define-XML files and generates Dataset-JSON files
conforming to the CDISC Dataset-JSON v1.1 specification.

It uses XSLT transformations via SaxonC to extract dataset metadata and structure
from Define-XML, then validates and outputs Dataset-JSON files for each dataset.

Key Features:
- Extracts dataset definitions from Define-XML
- Generates Dataset-JSON v1.1 compliant output files
- Schema validation against Dataset-JSON specification
- Command-line interface with configurable options
- Debug mode for troubleshooting
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from jsonschema import validate, ValidationError
from saxonche import PySaxonProcessor


class DatasetJSONGenerator:
    """
    Generator for Dataset-JSON files from Define-XML.
    
    This class processes CDISC Define-XML files and generates Dataset-JSON files
    for each dataset defined in the Define-XML using XSLT transformations.
    
    Attributes:
        define_file (str): Path to input Define-XML file
        output_dir (str): Directory for output Dataset-JSON files
        debug (bool): Enable debug mode with verbose output
        processor (PySaxonProcessor): Saxon XSLT processor instance
        schema (dict): Loaded Dataset-JSON schema
    """
    
    # Constants for schema and XSLT files
    SCHEMA_FILE = "dataset.schema.json"
    XSL_EXTRACT_FILE = "extract-list-ds.xsl"
    XSL_TRANSFORM_FILE = "transform-to-ds-json.xsl"
    
    def __init__(self, define_file, output_dir=None, debug=False):
        """
        Initialize the Dataset-JSON generator.
        
        Args:
            define_file (str): Path to Define-XML file
            output_dir (str, optional): Output directory for Dataset-JSON files
            debug (bool): Enable debug mode (default: False)
            
        Raises:
            FileNotFoundError: If required input files are not found
        """
        self.define_file = define_file
        self.output_dir = output_dir or os.path.dirname(define_file) or "."
        self.debug = debug
        
        # Validate input files exist
        self._validate_input_files()
        
        # Initialize Saxon processor
        self.processor = PySaxonProcessor(license=False)
        
        # Load schema
        self.schema = self._load_schema()
        
        # Statistics
        self.datasets_processed = 0
        self.datasets_failed = 0
        self.validation_errors = []
        self.generated_files = []

    def _validate_input_files(self):
        """
        Validate that all required input files exist.
        
        Raises:
            FileNotFoundError: If any required file is missing
        """
        required_files = {
            "Define-XML file": self.define_file,
            "Schema file": self.SCHEMA_FILE,
            "Extract XSL file": self.XSL_EXTRACT_FILE,
            "Transform XSL file": self.XSL_TRANSFORM_FILE
        }
        
        missing_files = []
        for file_type, file_path in required_files.items():
            if not os.path.exists(file_path):
                missing_files.append(f"{file_type}: {file_path}")
        
        if missing_files:
            raise FileNotFoundError(
                f"Required files not found:\n" + "\n".join(f"  - {f}" for f in missing_files)
            )

    def _load_schema(self):
        """
        Load and parse the Dataset-JSON schema file.
        
        Returns:
            dict: Parsed JSON schema
            
        Raises:
            json.JSONDecodeError: If schema file is invalid JSON
        """
        try:
            with open(self.SCHEMA_FILE, "r", encoding="utf-8") as f:
                schema = json.load(f)
            
            if self.debug:
                print(f"✓ Loaded schema from: {self.SCHEMA_FILE}")
            
            return schema
            
        except json.JSONDecodeError as e:
            raise json.JSONDecodeError(
                f"Invalid JSON in schema file {self.SCHEMA_FILE}: {e.msg}",
                e.doc, e.pos
            )

    def _extract_dataset_list(self):
        """
        Extract list of dataset names from Define-XML.
        
        Returns:
            list: List of dataset names
            
        Raises:
            Exception: If XSLT transformation fails
        """
        try:
            executable = self.processor.new_xslt30_processor().compile_stylesheet(
                stylesheet_file=self.XSL_EXTRACT_FILE
            )
            
            result = executable.transform_to_string(
                xdm_node=self.processor.parse_xml(xml_file_name=self.define_file)
            )

            if not result or not result.strip():
                return []

            datasets = [ds.strip() for ds in result.split(",") if ds.strip()]
            
            if self.debug:
                print(f"\n✓ Extracted {len(datasets)} datasets: {', '.join(datasets)}")
            
            return datasets
            
        except Exception as e:
            raise Exception(f"Failed to extract dataset list: {str(e)}")

    def _generate_dataset_json(self, dataset_name):
        """
        Generate Dataset-JSON for a single dataset.
        
        Args:
            dataset_name (str): Name of the dataset to generate
            
        Returns:
            dict: Generated Dataset-JSON data, or None if generation failed
        """
        try:
            executable_ds = self.processor.new_xslt30_processor().compile_stylesheet(
                stylesheet_file=self.XSL_TRANSFORM_FILE
            )
            
            executable_ds.set_parameter("dsName", self.processor.make_string_value(dataset_name))
            executable_ds.set_parameter(
                "datasetJSONCreationDateTime",
                self.processor.make_string_value(datetime.now().strftime('%Y-%m-%dT%H:%M:%S'))
            )
            
            result = executable_ds.transform_to_string(
                xdm_node=self.processor.parse_xml(xml_file_name=self.define_file)
            )

            if not result:
                print(f"  ✗ XSLT returned empty result for {dataset_name}")
                return None

            json_data = json.loads(result)
            
            if self.debug:
                print(f"  ✓ Generated JSON for dataset: {dataset_name}")
            
            return json_data
                
        except json.JSONDecodeError as e:
            print(f"  ✗ JSON decode error for {dataset_name}: {e.msg}")
            if self.debug:
                preview = repr(result[:200]) if result else "<None>"
                print(f"    First 200 chars of XSLT output: {preview}")
            return None
        except Exception as e:
            print(f"  ✗ Failed to generate JSON for {dataset_name}: {str(e)}")
            return None

    def _validate_dataset_json(self, dataset_name, json_data):
        """
        Validate Dataset-JSON against schema.
        
        Args:
            dataset_name (str): Name of the dataset
            json_data (dict): Dataset-JSON data to validate
            
        Returns:
            bool: True if validation passed, False otherwise
        """
        try:
            validate(json_data, self.schema)
            
            if self.debug:
                print(f"  ✓ Validation passed for: {dataset_name}")
            
            return True
            
        except ValidationError as e:
            error_msg = f"{dataset_name}: {e.message}"
            self.validation_errors.append(error_msg)
            print(f"  ✗ Validation failed for {dataset_name}: {e.message}")
            
            if self.debug and e.path:
                print(f"    Path: {' > '.join(str(p) for p in e.path)}")
            
            return False

    def _save_dataset_json(self, dataset_name, json_data):
        """
        Save Dataset-JSON to file.
        
        Args:
            dataset_name (str): Name of the dataset
            json_data (dict): Dataset-JSON data to save
            
        Returns:
            str: Path to saved file, or None if save failed
        """
        try:
            output_path = os.path.join(self.output_dir, f"{dataset_name}.json")
            
            # Create output directory if it doesn't exist
            os.makedirs(self.output_dir, exist_ok=True)
            
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(json_data, f, indent=4, ensure_ascii=False)
            
            self.generated_files.append(output_path)
            
            if self.debug:
                print(f"  ✓ Saved to: {output_path}")
            
            return output_path
            
        except Exception as e:
            print(f"  ✗ Failed to save {dataset_name}.json: {str(e)}")
            return None

    def process(self):
        """
        Main processing method to generate all Dataset-JSON files.
        
        Execution flow:
        1. Extract dataset list from Define-XML
        2. For each dataset:
           - Generate Dataset-JSON
           - Validate against schema
           - Save to output file
        3. Print summary statistics
        
        Returns:
            bool: True if all datasets processed successfully, False otherwise
        """
        print(f"\n{'='*70}")
        print(f"Dataset-JSON Generator")
        print(f"{'='*70}")
        print(f"Input Define-XML: {self.define_file}")
        print(f"Output directory: {self.output_dir}")
        print(f"{'='*70}\n")
        
        try:
            # Extract dataset list
            datasets = self._extract_dataset_list()
            
            if not datasets:
                print("⚠ No datasets found in Define-XML file")
                return False
            
            # Process each dataset
            print(f"\nProcessing {len(datasets)} dataset(s)...\n")
            
            for dataset_name in datasets:
                print(f"Processing: {dataset_name}")
                
                # Generate Dataset-JSON
                json_data = self._generate_dataset_json(dataset_name)
                if json_data is None:
                    self.datasets_failed += 1
                    continue
                
                # Validate
                if not self._validate_dataset_json(dataset_name, json_data):
                    self.datasets_failed += 1
                    # Still save even if validation failed
                
                # Save
                output_path = self._save_dataset_json(dataset_name, json_data)
                if output_path:
                    self.datasets_processed += 1
                else:
                    self.datasets_failed += 1
                
                print()  # Blank line between datasets
            
            # Print summary
            self._print_summary()
            
            return self.datasets_failed == 0
            
        except Exception as e:
            print(f"\n✗ Fatal error: {str(e)}")
            if self.debug:
                import traceback
                traceback.print_exc()
            return False

    def _print_summary(self):
        """Print processing summary statistics."""
        print(f"\n{'='*70}")
        print(f"Processing Summary")
        print(f"{'='*70}")
        print(f"Total datasets processed: {self.datasets_processed}")
        print(f"Failed datasets: {self.datasets_failed}")
        
        if self.validation_errors:
            print(f"\nValidation Errors ({len(self.validation_errors)}):")
            for error in self.validation_errors:
                print(f"  - {error}")
        
        if self.datasets_failed == 0:
            print(f"\n✅ All datasets processed successfully!")
        else:
            print(f"\n⚠ {self.datasets_failed} dataset(s) failed processing")
        
        print(f"{'='*70}\n")

    def validate_against_schema(self, schema_file, excel_output=None):
        """
        Validate all generated Dataset-JSON files against a JSON schema.

        Iterates over every file produced by the current run, validates each
        against the supplied JSON schema, and optionally writes a structured
        Excel report.  Falls back to _basic_schema_validation() when the
        schema file cannot be loaded.

        Args:
            schema_file (str): Path to the JSON schema file
            excel_output (str, optional): Path to output Excel validation report

        Returns:
            bool: True if all files pass validation, False otherwise
        """
        print(f"\nValidating output against schema: {schema_file}")

        try:
            with open(schema_file, "r", encoding="utf-8") as f:
                schema = json.load(f)

            print(f"✓ Schema file loaded successfully")

            validation_errors = []

            for file_path in self.generated_files:
                dataset_name = Path(file_path).stem
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        json_data = json.load(f)

                    validate(json_data, schema)

                    if self.debug:
                        print(f"  ✓ {dataset_name}: valid")

                except ValidationError as e:
                    error_msg = f"{dataset_name}: {e.message}"
                    validation_errors.append(error_msg)
                    print(f"  ✗ {dataset_name}: {e.message}")
                    if self.debug and e.path:
                        print(f"    Path: {' > '.join(str(p) for p in e.path)}")

                except Exception as e:
                    error_msg = f"{dataset_name}: {str(e)}"
                    validation_errors.append(error_msg)
                    print(f"  ✗ {dataset_name}: {str(e)}")

            if not validation_errors:
                print("✅ All datasets validated successfully!")
                if excel_output:
                    self._write_validation_excel([], excel_output, passed=True)
                return True
            else:
                print(f"⚠️  Validation found {len(validation_errors)} issue(s)")
                if excel_output:
                    self._write_validation_excel(validation_errors, excel_output, passed=False)
                    print(f"   ✓ Validation report written to: {excel_output}")
                else:
                    for error in validation_errors[:10]:
                        print(f"   • {error}")
                    if len(validation_errors) > 10:
                        print(f"   ... and {len(validation_errors) - 10} more errors")
                print()
                return False

        except FileNotFoundError:
            print(f"❌ Schema file not found: {schema_file}")
            return False
        except json.JSONDecodeError as e:
            print(f"❌ Error parsing schema file {schema_file}: {e}")
            return self._basic_schema_validation(None)
        except Exception as e:
            print(f"❌ Validation error: {e}")
            if self.debug:
                import traceback
                traceback.print_exc()
            return False

    def _basic_schema_validation(self, schema):
        """
        Perform basic structural validation on generated Dataset-JSON files.

        Checks that each generated file contains the required root-level fields
        and that each column entry has mandatory attributes.  Used as a fallback
        when the JSON schema file cannot be loaded.

        Args:
            schema: Unused – kept for API consistency with create_define_json.py

        Returns:
            bool: True if all files pass basic checks, False otherwise
        """
        errors = []

        print("Performing basic structural checks...\n")

        required_root_fields = [
            "datasetJSONCreationDateTime",
            "datasetJSONVersion",
            "itemGroupOID",
            "records",
            "name",
            "label",
            "columns",
        ]
        required_column_fields = ["itemOID", "name", "label", "dataType"]

        for file_path in self.generated_files:
            dataset_name = Path(file_path).stem
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    json_data = json.load(f)

                for field in required_root_fields:
                    if field not in json_data:
                        errors.append(f"{dataset_name}: Missing required field: {field}")

                for idx, col in enumerate(json_data.get("columns", [])):
                    for req in required_column_fields:
                        if req not in col:
                            errors.append(f"{dataset_name}: columns[{idx}]: Missing {req}")

            except Exception as e:
                errors.append(f"{dataset_name}: Could not read file: {e}")

        if errors:
            print("❌ VALIDATION ERRORS:")
            for error in errors:
                print(f"   • {error}")
            print()
        else:
            print("✅ Basic structural validation passed!")

        print()
        return len(errors) == 0

    def _write_validation_excel(self, validation_errors, excel_path, passed=True):
        """
        Write validation results to an Excel file.

        Produces a two-sheet workbook: a Summary sheet with run statistics
        and a Validation Errors sheet with one row per error.

        Args:
            validation_errors (list): List of validation error message strings
            excel_path (str): Destination path for the Excel workbook
            passed (bool): Overall pass/fail status (default: True)

        Returns:
            bool: True if the file was written successfully, False otherwise
        """
        try:
            import pandas as pd
            from openpyxl.styles import Font, PatternFill, Alignment

            summary_data = {
                "Validation Date": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                "Schema File": [self.SCHEMA_FILE],
                "Output Directory": [self.output_dir],
                "Status": ["PASSED" if passed else "FAILED"],
                "Total Errors": [len(validation_errors)],
                "Datasets Processed": [self.datasets_processed],
                "Datasets Failed": [self.datasets_failed],
            }

            if validation_errors:
                errors_data = []
                for idx, error in enumerate(validation_errors, 1):
                    parts = error.split(": ", 1)
                    dataset = parts[0] if len(parts) >= 2 else "N/A"
                    issue = parts[1] if len(parts) >= 2 else error
                    errors_data.append({
                        "Error #": idx,
                        "Dataset": dataset,
                        "Issue": issue,
                        "Full Error": error,
                    })
                errors_df = pd.DataFrame(errors_data)
            else:
                errors_df = pd.DataFrame({"Message": ["No validation errors found"]})

            summary_df = pd.DataFrame(summary_data)

            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                errors_df.to_excel(writer, sheet_name="Validation Errors", index=False)

                # Format Summary sheet
                summary_ws = writer.sheets["Summary"]
                for cell in summary_ws[1]:
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")

                # Format Errors sheet
                if validation_errors:
                    errors_ws = writer.sheets["Validation Errors"]
                    for cell in errors_ws[1]:
                        cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")

                    errors_ws.column_dimensions["A"].width = 8    # Error #
                    errors_ws.column_dimensions["B"].width = 20   # Dataset
                    errors_ws.column_dimensions["C"].width = 60   # Issue
                    errors_ws.column_dimensions["D"].width = 15   # Full Error (hidden)
                    errors_ws.column_dimensions["D"].hidden = True

                    for row in errors_ws.iter_rows(
                        min_row=2, max_row=len(validation_errors) + 1, min_col=3, max_col=3
                    ):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Auto-adjust Summary sheet column widths
                for column in summary_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    summary_ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            return True

        except ImportError:
            print("   ⚠️  pandas and openpyxl required for Excel output")
            print("   Install with: pip install pandas openpyxl")
            return False
        except Exception as e:
            print(f"   ⚠️  Error writing Excel file: {e}")
            return False


def main():
    """
    Main entry point for command-line execution.
    
    Parses command-line arguments and orchestrates the Dataset-JSON
    generation process.
    """
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    parser = argparse.ArgumentParser(
        description="Generate Dataset-JSON files from CDISC Define-XML.",
        epilog="Example: python create-shells-ds-json.py --define_file define-2-1-ADaM.xml --output_dir output/"
    )
    
    parser.add_argument(
        "--define_file",
        required=True,
        help="Path to Define-XML file"
    )
    parser.add_argument(
        "--output_dir",
        help="Output directory for Dataset-JSON files (default: same as define_file)"
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable debug mode with verbose output"
    )
    parser.add_argument(
        "--validate",
        nargs="?",
        const=DatasetJSONGenerator.SCHEMA_FILE,
        default=None,
        help=f"Validate generated Dataset-JSON files against a JSON schema "
             f"(default: {DatasetJSONGenerator.SCHEMA_FILE})"
    )
    parser.add_argument(
        "--validation_report",
        required=False,
        help="Path to Excel file for validation report (e.g., validation_report.xlsx)"
    )
    
    args = parser.parse_args()

    # Validate conditional requirements
    if args.validate and not args.validation_report:
        parser.error("--validation_report is required when --validate is used")
    
    try:
        generator = DatasetJSONGenerator(
            define_file=args.define_file,
            output_dir=args.output_dir,
            debug=args.debug
        )
        
        success = generator.process()

        print(f"\n✅ Dataset-JSON files created successfully in: {generator.output_dir}")

        # Validate if requested
        if args.validate:
            is_valid = generator.validate_against_schema(
                args.validate, excel_output=args.validation_report
            )
            if not is_valid:
                exit(1)
        
        # Exit with appropriate code
        exit(0 if success else 1)
        
    except FileNotFoundError as e:
        print(f"\n✗ Error: {str(e)}")
        exit(1)
    except Exception as e:
        print(f"\n✗ Unexpected error: {str(e)}")
        if args.debug:
            import traceback
            traceback.print_exc()
        exit(1)


if __name__ == "__main__":
    main()
